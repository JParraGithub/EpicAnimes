               

from datetime import timedelta

from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from difflib import SequenceMatcher

import json

import unicodedata
from urllib.parse import urlsplit



from django.conf import settings

import logging

from django.contrib import messages

from django.contrib.auth import login as auth_login

from django.contrib.auth.decorators import login_required

from django.contrib.auth.models import User, Group

from django.contrib.auth.views import LoginView, PasswordResetView

from django.core.exceptions import ValidationError

from django.core.mail import EmailMultiAlternatives, send_mail

from django.core.validators import validate_email

from django.core.cache import cache
from django.core.files.images import get_image_dimensions

from django.contrib.auth.password_validation import validate_password

from django.db import transaction, connection

from django.db.models import Sum, F, Count, Exists, OuterRef, Q

from django.http import (

    JsonResponse,

    HttpResponseBadRequest,

    HttpResponseForbidden,

    HttpResponse,

    HttpResponseNotAllowed,

)

from django.shortcuts import render, redirect, get_object_or_404

from django.templatetags.static import static
from django.template.loader import render_to_string

from django.urls import reverse, reverse_lazy

from django.utils import timezone

from django.utils.dateparse import parse_date
from django.utils.html import strip_tags

from django.utils.http import url_has_allowed_host_and_scheme

from django.views.decorators.http import require_http_methods



from .forms import LoginForm, RegistroClienteForm, PostulacionVendedorForm, TwoFactorLoginForm
from .models import Vendedor, Producto, Venta, Compra, PerfilCliente, PostulacionVendedor, NewsletterSubscriber
from .payments import (
    paypal_capture_order,
    paypal_create_order,
    PayPalError,
    paypal_is_configured,
    paypal_amount_step,
    paypal_conversion_summary,
    get_paypal_conversion_rate,
    normalize_paypal_totals,
)
from .chatbot import responder as chatbot_responder

logger = logging.getLogger(__name__)

PRODUCT_IMAGE_MAX_MB = 2
PRODUCT_IMAGE_MAX_WIDTH = 1200
PRODUCT_IMAGE_MAX_HEIGHT = 1200


def _validar_imagen_producto(imagen, *, max_mb=PRODUCT_IMAGE_MAX_MB, max_width=PRODUCT_IMAGE_MAX_WIDTH, max_height=PRODUCT_IMAGE_MAX_HEIGHT):
    """Verifica el tamaño en MB y las dimensiones máximas permitidas para las imágenes de productos."""
    if not imagen:
        return
    max_bytes = int(max_mb * 1024 * 1024)
    if getattr(imagen, "size", 0) > max_bytes:
        raise ValidationError(f"La imagen es muy pesada. Usa archivos de hasta {max_mb} MB.")
    try:
        imagen.seek(0)
    except Exception:
        pass
    width, height = get_image_dimensions(imagen)
    try:
        imagen.seek(0)
    except Exception:
        pass
    if not width or not height:
        raise ValidationError("No pude leer la imagen. Usa formatos JPG o PNG válidos.")
    if width > max_width or height > max_height:
        raise ValidationError(f"La imagen debe medir como máximo {max_width}x{max_height} px.")


@require_http_methods(["POST"])

def newsletter_suscribir(request):
    """Suscribe un correo al boletín y notifica el resultado."""
    email = (request.POST.get("newsletter_email") or "").strip()
    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or "index"
    anchor = (request.POST.get("next_anchor") or request.POST.get("anchor") or "").strip().lstrip("#")
    segmento = (request.POST.get("segmento") or "").strip()
    source_hint = (request.POST.get("source_path") or request.POST.get("next") or request.META.get("HTTP_REFERER") or "").strip()
    wants_json = (
        "application/json" in (request.headers.get("Accept") or "").lower()
        or request.headers.get("HX-Request")
        or request.headers.get("X-Requested-With") == "XMLHttpRequest"
    )

    def _normalize_source(value: str) -> str:
        if not value:
            return ""
        try:
            parts = urlsplit(value)
        except ValueError:
            return value[:255]
        path_value = parts.path or "/"
        if parts.query:
            path_value = f"{path_value}?{parts.query}"
        return path_value[:255]

    source_path = _normalize_source(source_hint)
    segmento_value = segmento[:60]

    def _redirect_default():
        target = None
        if next_url and url_has_allowed_host_and_scheme(
            next_url,
            allowed_hosts={request.get_host()},
            require_https=request.is_secure(),
        ):
            target = next_url
        if not target:
            target = reverse("index")
        if anchor:
            base = target.split("#", 1)[0]
            target = f"{base}#{anchor}"
        return redirect(target)

    def _respond(level: str, text: str, *, created: bool = False):
        if wants_json:
            status_code = 201 if created else (200 if level != "error" else 400)
            return JsonResponse({"ok": level != "error", "created": created, "message": text}, status=status_code)
        notifier = {
            "error": messages.error,
            "success": messages.success,
            "warning": messages.warning,
            "info": messages.info,
        }.get(level, messages.info)
        notifier(request, text)
        return _redirect_default()

    if not email:
        return _respond("error", "Ingresa tu correo para suscribirte.")

    try:
        validate_email(email)
    except ValidationError:
        return _respond("error", "El correo ingresado no es válido.")

    email = email.lower()
    defaults = {}
    if segmento_value:
        defaults["segmento"] = segmento_value
    if source_path:
        defaults["ruta_origen"] = source_path

    subscriber, created = NewsletterSubscriber.objects.get_or_create(email=email, defaults=defaults)

    if created:
        asunto = "¡Bienvenido a la comunidad EpicAnimes!"
        segment_label = segmento_value or "la comunidad EpicAnimes"
        cta_url = request.build_absolute_uri(reverse("contacto"))
        html_body = render_to_string(
            "emails/newsletter_welcome.html",
            {
                "cta_url": cta_url,
                "segment_label": segment_label,
                "support_email": settings.EMAIL_HOST_USER,
            },
        )
        text_body = strip_tags(html_body)

        try:
            message = EmailMultiAlternatives(
                subject=asunto,
                body=text_body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[email],
            )
            message.attach_alternative(html_body, "text/html")
            message.send(fail_silently=False)
            return _respond("success", "Gracias por unirte a la comunidad. Te enviamos un correo de bienvenida.", created=True)
        except Exception:
            logger.exception("Fallo al enviar correo de bienvenida a %s", email)
            return _respond("warning", "Te suscribimos, pero no pudimos enviar el correo de bienvenida.", created=True)

    updated_fields = []
    if segmento_value:
        existing = [seg.strip() for seg in (subscriber.segmento or "").split(",") if seg.strip()]
        existing_lower = {seg.lower() for seg in existing}
        if segmento_value.lower() not in existing_lower:
            existing.append(segmento_value)
            subscriber.segmento = ", ".join(existing)
            updated_fields.append("segmento")
    if source_path and source_path != subscriber.ruta_origen:
        subscriber.ruta_origen = source_path
        updated_fields.append("ruta_origen")
    if updated_fields:
        subscriber.save(update_fields=updated_fields)
        return _respond("info", "Actualizamos tus preferencias de newsletter.")

    return _respond("info", "Ese correo ya está suscrito a nuestras noticias.")

@require_http_methods(["POST"])
def api_chatbot_ask(request):
    """Responde preguntas enviadas al chatbot con mensajes JSON."""
    rol_usuario = obtener_rol_usuario(request.user)
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError, AttributeError, TypeError):
        return JsonResponse({"ok": False, "error": "No pude entender el formato de la solicitud."}, status=400)

    question = (payload.get("message") or "").strip()
    if not question:
        return JsonResponse({"ok": False, "error": "Por favor escribe una pregunta para que pueda ayudarte."}, status=400)

    try:
        result = chatbot_responder(question, rol_usuario)
    except RuntimeError as exc:
        logger.warning("Chatbot temporalmente inhabilitado: %s", exc)
        return JsonResponse({"ok": False, "error": str(exc)}, status=503)
    except Exception:
        logger.exception("Error al procesar la pregunta del chatbot")
        return JsonResponse({"ok": False, "error": "Tu pregunta no pudo procesarse por ahora. Intenta nuevamente en un momento."}, status=500)

    answer = result.get("answer") or "No tengo una respuesta disponible en este momento."
    confidence = float(result.get("confidence", 0.0))
    return JsonResponse(
        {"ok": True, "answer": answer, "confidence": confidence, "user_role": rol_usuario}
    )

class CarritoError(Exception):

    """Error controlado al procesar el carrito."""





def obtener_rol_usuario(user):

    """Determina el rol actual del usuario autenticado."""
    if not user.is_authenticated:

        return "anonimo"

    if user.is_superuser or user.is_staff:

        return "administrador"

    if user.groups.filter(name="Vendedores").exists():

        return "vendedor"

    return "comprador"





def _get_cart(request):

    """Obtiene y limpia el carrito almacenado en la sesión."""
    raw = request.session.get("cart", {})

    cart = {}

    for key, value in raw.items():

        try:

            cantidad = int(value)

        except (TypeError, ValueError):

            continue

        if cantidad <= 0:

            continue

        cart[str(key)] = cantidad

    return cart





def _save_cart(request, cart):

    """Persiste el carrito en la sesión y marca la modificación."""
    request.session["cart"] = cart

    request.session.modified = True





def _cart_count(cart):

    """Cuenta la cantidad total de unidades registradas en el carrito."""
    return sum(cart.values())





def _resolver_datos_cliente(request, datos_cliente=None):

    """Compila y valida los datos de despacho del cliente."""
    datos_cliente = datos_cliente or {}

    nombre = (

        datos_cliente.get("nombre")

        or request.user.get_full_name()

        or request.user.username

        or ""

    ).strip()

    correo = (datos_cliente.get("email") or request.user.email or "").strip()

    telefono = (datos_cliente.get("telefono") or "").strip()

    direccion = (datos_cliente.get("direccion") or "").strip()

    ciudad = (datos_cliente.get("ciudad") or "").strip()

    notas = (datos_cliente.get("notas") or "").strip()



    if not nombre or not correo or not direccion or not ciudad:

        raise CarritoError("Completa los datos de envío antes de pagar.")



    datos = {

        "nombre": nombre,

        "correo": correo,

        "telefono": telefono,

        "direccion": direccion,

        "ciudad": ciudad,

        "notas": notas,

    }



    if request.user.is_authenticated:

        perfil, _ = PerfilCliente.objects.get_or_create(user=request.user)

        perfil.nombre = nombre

        perfil.email = correo

        perfil.telefono = telefono

        perfil.direccion = direccion

        perfil.ciudad = ciudad

        perfil.save()



    return datos





def _calcular_lineas_y_total(cart, *, lock=False):

    """Valida el stock de cada producto y acumula la suma del carrito."""
    if not cart:

        raise CarritoError("El carrito está vacío.")



    ids = [int(pk) for pk in cart.keys()]

    qs = (
        Producto.objects.filter(id__in=ids)
        .select_related("vendedor", "vendedor__usuario")
    )

    if lock:

        qs = qs.select_for_update()

    productos = list(qs)

    productos_map = {p.id: p for p in productos}

    if len(productos_map) != len(ids):

        raise CarritoError("Uno de los productos ya no está disponible.")



    lineas = []

    total = Decimal("0")

    for pid in ids:

        cantidad = cart[str(pid)]

        if cantidad <= 0:

            raise CarritoError("Cantidad inválida en el carrito.")

        producto = productos_map.get(pid)

        if producto is None:

            raise CarritoError("Producto no encontrado.")

        if producto.existencias < cantidad:

            raise CarritoError(f"Stock insuficiente para {producto.nombre}.")

        subtotal = producto.precio * cantidad

        total += subtotal

        lineas.append((producto, cantidad, subtotal))



    return lineas, total



def _correo_remitente_default():

    """Entrega el remitente base para los correos automáticos."""
    return (

        getattr(settings, "DEFAULT_FROM_EMAIL", None)

        or getattr(settings, "EMAIL_HOST_USER", None)

        or "no-reply@epicanimes.com"

    )



def _formatear_monto_correo(valor, paso):

    """Formatea montos en función de la resolución monetaria."""
    paso = paso or Decimal("0.01")

    cuantizado = valor.quantize(paso)

    if paso == Decimal("1"):

        return f"$ {int(cuantizado)}"

    return f"$ {cuantizado:.2f}"



def _resumen_productos_correo(lineas, paso):

    """Construye un resumen textual de los productos comprados."""
    partes = []

    for producto, cantidad, subtotal in lineas:

        partes.append(f"- {producto.nombre} x{cantidad} ({_formatear_monto_correo(subtotal, paso)})")

    return "\n".join(partes) if partes else "- Sin productos asociados"



def _enviar_correo_simple(asunto, cuerpo, destinatario):

    """Centraliza el envío y loguea fallos sin interrumpir la compra."""
    remitente = _correo_remitente_default()

    try:

        send_mail(asunto, cuerpo, remitente, [destinatario], fail_silently=False)

    except Exception:

        logger.exception("No se pudo enviar el correo '%s' a %s", asunto, destinatario)



def _notificar_actores_compra(datos_cliente, lineas, total, referencia_pago, paso_moneda):

    """Envía un correo al comprador y a cada vendedor involucrado."""
    if not lineas:

        return

    nombre_cliente = datos_cliente.get("nombre") or "Cliente EpicAnimes"

    correo_cliente = datos_cliente.get("correo") or datos_cliente.get("email")

    telefono = datos_cliente.get("telefono") or "No informado"

    direccion = datos_cliente.get("direccion") or "Sin dirección"

    ciudad = datos_cliente.get("ciudad") or ""

    notas = datos_cliente.get("notas") or "Sin notas"

    referencia = referencia_pago or "Sin referencia"

    total_formateado = _formatear_monto_correo(total, paso_moneda)

    detalle_general = _resumen_productos_correo(lineas, paso_moneda)

    if correo_cliente:

        cuerpo_cliente = (

            f"Hola {nombre_cliente},\n\n"

            "EpicAnimes confirma que su compra fue registrada con éxito.\n"

            f"Referencia: {referencia}\n"

            f"Total: {total_formateado}\n\n"

            "Detalle de productos:\n"

            f"{detalle_general}\n\n"

            "Datos de envío:\n"

            f"- Dirección: {direccion}\n"

            f"- Ciudad: {ciudad or 'No indicada'}\n"

            f"- Teléfono: {telefono}\n"

            f"- Notas: {notas}\n\n"

            "Gracias por confiar en EpicAnimes."

        )

        _enviar_correo_simple("EpicAnimes | Compra registrada", cuerpo_cliente, correo_cliente)

    vendedores = {}

    for producto, cantidad, subtotal in lineas:

        vendedor = getattr(producto, "vendedor", None)

        usuario_vendedor = getattr(vendedor, "usuario", None) if vendedor else None

        if not usuario_vendedor or not usuario_vendedor.email:

            continue

        vendedores.setdefault(usuario_vendedor, []).append((producto, cantidad, subtotal))

    for usuario_vendedor, productos_vendedor in vendedores.items():

        nombre_vendedor = (

            usuario_vendedor.get_full_name()

            or usuario_vendedor.username

            or "Vendedor EpicAnimes"

        )

        total_vendedor = sum((subtotal for _, _, subtotal in productos_vendedor), Decimal("0"))

        detalle_vendedor = _resumen_productos_correo(productos_vendedor, paso_moneda)

        cuerpo_vendedor = (

            f"Hola {nombre_vendedor},\n\n"

            "Se registró una nueva compra relacionada con tus productos en EpicAnimes.\n"

            f"Referencia: {referencia}\n"

            f"Total asociado: {_formatear_monto_correo(total_vendedor, paso_moneda)}\n\n"

            "Detalle de productos:\n"

            f"{detalle_vendedor}\n\n"

            "Datos de envío del comprador:\n"

            f"- Nombre: {nombre_cliente}\n"

            f"- Correo: {correo_cliente or 'Sin correo'}\n"

            f"- Teléfono: {telefono}\n"

            f"- Dirección: {direccion}\n"

            f"- Ciudad: {ciudad or 'No indicada'}\n"

            f"- Notas: {notas}\n\n"

            "Prepara el despacho y actualiza el seguimiento dentro del panel de vendedor.\n"

            "Equipo EpicAnimes"

        )

        _enviar_correo_simple("EpicAnimes | Nueva orden para despacho", cuerpo_vendedor, usuario_vendedor.email)



def _build_cart_items(cart):

    """Convierte el contenido del carrito en objetos enriquecidos y calcula totales."""
    if not cart:

        return [], Decimal("0"), True

    producto_ids = [int(pid) for pid in cart.keys()]

    productos = (

        Producto.objects.select_related("vendedor", "vendedor__usuario")

        .filter(id__in=producto_ids)

    )

    productos_map = {p.id: p for p in productos}

    default_image_url = static("images/Imagen1.png")

    items = []

    total = Decimal("0")

    for producto in productos:

        cantidad = cart.get(str(producto.id), 0)

        subtotal = producto.precio * cantidad

        total += subtotal

        items.append(

            {

                "producto": producto,

                "cantidad": cantidad,

                "subtotal": subtotal,

                "sin_stock": producto.existencias < cantidad,

                "imagen_url": producto.imagen.url if getattr(producto, "imagen", None) else default_image_url,

            }

        )

    items.sort(key=lambda x: x["producto"].nombre.lower())

    total = total.quantize(Decimal("0.01")) if items else Decimal("0.00")

    puede_pagar = bool(items) and all(not item["sin_stock"] for item in items)

    return items, total, puede_pagar





def _normalize_text(text):

    """Normaliza cadenas para comparaciones insensibles a tildes."""
    if not text:

        return ""

    normalized = unicodedata.normalize("NFKD", text)

    stripped = ''.join(ch for ch in normalized if not unicodedata.combining(ch))

    return stripped.lower().strip()





def _smart_tokenize(text):

    """Tokeniza una cadena manteniendo una versión completa normalizada."""
    base = _normalize_text(text)

    tokens = [tok for tok in base.split() if tok]

    if base and base not in tokens:

        tokens.append(base)

    return base, tokens





def _smart_match_score(producto, query_tokens, query_full):

    """Calcula un puntaje heurístico entre un producto y una consulta."""
    campos = [

        producto.nombre,

        producto.marca,

        producto.categoria,

        producto.descripcion,

    ]

    campos_norm = [_normalize_text(campo) for campo in campos]



    mejor = 0.0

    coincidencias = False

    tokens = query_tokens or ([query_full] if query_full else [])



    for campo in campos_norm:

        if not campo:

            continue

        for token in tokens:

            if not token:

                continue

            if token in campo:

                coincidencias = True

                mejor = max(mejor, min(1.0, 0.75 + len(token) / max(len(campo), len(token)) * 0.25))

            else:

                ratio = SequenceMatcher(None, campo, token).ratio()

                if ratio >= 0.62:

                    coincidencias = True

                    mejor = max(mejor, ratio)

        if query_full:

            ratio_full = SequenceMatcher(None, campo, query_full).ratio()

            if ratio_full >= 0.65:

                coincidencias = True

                mejor = max(mejor, ratio_full)



    return coincidencias, mejor



                                                              

                                      

                                                              



@require_http_methods(["GET", "POST"])

def VistaIndex(request):

    """Construye la página principal con listados y destacados."""
    rol_usuario = obtener_rol_usuario(request.user)



    cart = _get_cart(request)

    cart_count = _cart_count(cart)



    filtro_busqueda = (request.GET.get("q") or "").strip()
    filtro_categoria = (request.GET.get("categoria") or "").strip()
    filtro_marca = (request.GET.get("marca") or "").strip()
    filtro_calidad = (request.GET.get("calidad") or "").strip()
    filtro_precio_min_raw = (request.GET.get("precio_min") or "").strip()
    filtro_precio_max_raw = (request.GET.get("precio_max") or "").strip()
    filtro_en_stock = (request.GET.get("en_stock") or "").strip()
    filtro_orden = (request.GET.get("orden") or "recientes").strip()



    productos_qs = Producto.objects.select_related("vendedor", "vendedor__usuario")

    if filtro_categoria and filtro_categoria.lower() != "todos":
        productos_qs = productos_qs.filter(categoria__iexact=filtro_categoria)
    if filtro_marca and filtro_marca.lower() != "todas":
        productos_qs = productos_qs.filter(marca__iexact=filtro_marca)
    if filtro_calidad and filtro_calidad.lower() != "todas":
        productos_qs = productos_qs.filter(calidad__iexact=filtro_calidad)
                                                                 
    try:
        if filtro_precio_min_raw:
            productos_qs = productos_qs.filter(precio__gte=Decimal(filtro_precio_min_raw))
    except (InvalidOperation, ValueError):
        pass
    try:
        if filtro_precio_max_raw:
            productos_qs = productos_qs.filter(precio__lte=Decimal(filtro_precio_max_raw))
    except (InvalidOperation, ValueError):
        pass
    if filtro_en_stock in ("1", "on", "true", "True"):
        productos_qs = productos_qs.filter(existencias__gt=0)



    productos_qs = productos_qs.order_by("-fecha_ingreso", "nombre")

    productos = list(productos_qs)



    if filtro_busqueda:

        query_full_norm, query_tokens = _smart_tokenize(filtro_busqueda)

        filtrados = []

        for prod in productos:

            coincide, score = _smart_match_score(prod, query_tokens, query_full_norm)

            if coincide:

                prod._search_score = score

                filtrados.append(prod)

        productos = filtrados

    else:

        for prod in productos:

            prod._search_score = 1.0



    if filtro_orden == "precio_asc":

        productos.sort(key=lambda p: float(p.precio or 0))

    elif filtro_orden == "precio_desc":

        productos.sort(key=lambda p: float(p.precio or 0), reverse=True)

    elif filtro_orden == "stock":

        productos.sort(key=lambda p: p.existencias or 0, reverse=True)

    else:

        if filtro_busqueda:

            productos.sort(key=lambda p: getattr(p, "_search_score", 0), reverse=True)



    default_image_url = static("images/Imagen1.png")
    now_date = timezone.now().date()
    for prod in productos:
        prod.imagen_url = prod.imagen.url if getattr(prod, "imagen", None) else default_image_url
        try:
            prod.is_new = (now_date - (prod.fecha_ingreso or now_date)).days <= 14
        except Exception:
            prod.is_new = False

    categorias = (
        Producto.objects.exclude(categoria__isnull=True)
        .exclude(categoria__exact="")
        .order_by("categoria")
        .values_list("categoria", flat=True)
        .distinct()
    )
    marcas = (
        Producto.objects.exclude(marca__isnull=True)
        .exclude(marca__exact="")
        .order_by("marca")
        .values_list("marca", flat=True)
        .distinct()
    )
    calidades = (
        Producto.objects.exclude(calidad__isnull=True)
        .exclude(calidad__exact="")
        .order_by("calidad")
        .values_list("calidad", flat=True)
        .distinct()
    )

                                                                               
    sugerencias_busqueda = []
    try:
        sugerencias_busqueda.extend(list(categorias[:8]))
    except Exception:
        pass
    try:
        sugerencias_busqueda.extend(list(marcas[:8]))
    except Exception:
        pass
    try:
        nombres_recientes = list(
            Producto.objects.order_by("-fecha_ingreso").values_list("nombre", flat=True)[:10]
        )
        sugerencias_busqueda.extend(nombres_recientes)
    except Exception:
        pass



    login_form = LoginForm(request)

    registro_form = RegistroClienteForm()



    if request.method == "POST":

        action = request.POST.get("form_type")

        if action == "login":

            login_form = LoginForm(request, data=request.POST)

            if login_form.is_valid():

                user = login_form.get_user()

                auth_login(request, user)

                destino = request.POST.get("next") or ""

                if not destino:

                    rol_destino = obtener_rol_usuario(user)

                    if rol_destino == "administrador":

                        destino = reverse("dashboard_administrador")

                    elif rol_destino == "vendedor":

                        destino = reverse("dashboard_vendedor")

                    else:

                        destino = reverse("index")

                messages.success(request, "Sesión iniciada correctamente.")

                return redirect(destino)

            messages.error(request, "No pudimos iniciar la sesión. Revisa tus datos.")

        elif action == "register":

            registro_form = RegistroClienteForm(request.POST)

            if registro_form.is_valid():

                with transaction.atomic():

                    user = registro_form.save()

                    grupo_clientes, _ = Group.objects.get_or_create(name="Clientes")

                    user.groups.add(grupo_clientes)

                # Notificar por correo que la cuenta fue creada
                try:
                    if user.email:
                        asunto = "¡Bienvenido a EpicAnimes!"
                        cta_url = request.build_absolute_uri(reverse("index"))
                        html_body = render_to_string(
                            "emails/signup_welcome.html",
                            {"username": user.username, "cta_url": cta_url, "support_email": settings.EMAIL_HOST_USER},
                        )
                        text_body = strip_tags(html_body)
                        msg = EmailMultiAlternatives(asunto, text_body, settings.DEFAULT_FROM_EMAIL, [user.email])
                        msg.attach_alternative(html_body, "text/html")
                        msg.send(fail_silently=True)
                except Exception:
                    logger.exception("No se pudo enviar correo de bienvenida tras registro (user=%s)", user.id)

                auth_login(request, user)

                messages.success(request, "Cuenta creada. Ya puedes comprar.")

                return redirect("index")

            messages.error(request, "Revisa los datos del formulario de registro.")

        else:

            messages.error(request, "Acción no reconocida.")



    contexto = {

        "productos": productos,
        "categorias": categorias,
        "marcas": marcas,
        "calidades": calidades,

        "rol_usuario": rol_usuario,

        "login_form": login_form,

        "registro_form": registro_form,

        "filtros": {

            "q": filtro_busqueda,
            "categoria": filtro_categoria,
            "marca": filtro_marca,
            "calidad": filtro_calidad,
            "precio_min": filtro_precio_min_raw,
            "precio_max": filtro_precio_max_raw,
            "en_stock": filtro_en_stock in ("1", "on", "true", "True"),
            "orden": filtro_orden,

        },

        "cart_count": cart_count,

        "puede_comprar": rol_usuario == "comprador",

        "sugerencias_busqueda": sugerencias_busqueda,

    }

    return render(request, "public/index.html", contexto)





@require_http_methods(["GET", "POST"])

def VistaContacto(request):

    """Renderiza la página de contacto con datos de soporte."""
    cart = _get_cart(request)

    cart_count = _cart_count(cart)

    rol_usuario = obtener_rol_usuario(request.user)

    form = PostulacionVendedorForm(request.POST or None)

    if request.method == "POST":

        if form.is_valid():

            form.save()

            messages.success(request, "Gracias por postular. Te contactaremos pronto.")

            return redirect("contacto")

        else:

            messages.error(request, "Revisa los datos del formulario.")

    contexto = {

        "form": form,

        "cart_count": cart_count,

        "rol_usuario": rol_usuario,

    }

    return render(request, "public/contacto.html", contexto)





def VistaProductoDetalle(request, producto_id):

    """Presenta el detalle de un producto específico."""
    producto = get_object_or_404(

        Producto.objects.select_related("vendedor", "vendedor__usuario"),

        pk=producto_id,

    )

    rol_usuario = obtener_rol_usuario(request.user)

    cart = _get_cart(request)

    cart_count = _cart_count(cart)

    default_image_url = static("images/Imagen1.png")



    producto.imagen_url = producto.imagen.url if getattr(producto, "imagen", None) else default_image_url

    vendedor_nombre = None

    if producto.vendedor and producto.vendedor.usuario:

        vendedor_nombre = producto.vendedor.usuario.get_full_name() or producto.vendedor.usuario.username



    relacionados_qs = (

        Producto.objects.exclude(pk=producto.pk)

        .filter(categoria__iexact=producto.categoria)

        .order_by("-fecha_ingreso")[:4]

    )

    relacionados = list(relacionados_qs)

    for rel in relacionados:

        rel.imagen_url = rel.imagen.url if getattr(rel, "imagen", None) else default_image_url



    contexto = {

        "producto": producto,

        "vendedor_nombre": vendedor_nombre,

        "relacionados": relacionados,

        "rol_usuario": rol_usuario,

        "puede_comprar": rol_usuario == "comprador",

        "cart_count": cart_count,

    }

    return render(request, "public/producto_detalle.html", contexto)





@require_http_methods(["GET", "POST"])

def VistaRegistro(request):

    """

    Registro simple de usuarios (self-service).

    Crea un usuario activo y redirige al login.

    """

    if request.method == "POST":

        username = (request.POST.get("username") or "").strip()
        email = (request.POST.get("email") or "").strip()
        p1 = (request.POST.get("password1") or "").strip()
        p2 = (request.POST.get("password2") or "").strip()
        terms_raw = (request.POST.get("terms") or "").strip().lower()



        errores = []

        if not username:
            errores.append("El nombre de usuario es obligatorio.")

        if not email:
            errores.append("El correo electrónico es obligatorio.")
        else:
            try:
                validate_email(email)
            except ValidationError:
                errores.append("Ingresa un correo electrónico válido.")

        if email and User.objects.filter(email__iexact=email).exists():
            errores.append("El correo electrónico ya está registrado.")

        if not p1:
            errores.append("La contraseña es obligatoria.")

        if p1 and len(p1) < 6:
            errores.append("La contraseña debe tener al menos 6 caracteres.")

        if p1 != p2:
            errores.append("Las contraseñas no coinciden.")

        if User.objects.filter(username=username).exists():
            errores.append("El nombre de usuario ya existe.")

        # Validación de aceptación de términos
        if terms_raw not in ("on", "1", "true", "sí", "si"): 
            errores.append("Debes aceptar los términos y condiciones.")



        if errores:

            for e in errores:

                messages.error(request, e)

                                            

            return render(request, "registration/signup.html", {

                "form": {"username": username, "email": email}

            })



        u = User.objects.create_user(username=username, email=email, password=p1)

        u.is_active = True

        u.save()

        grupo_clientes, _ = Group.objects.get_or_create(name="Clientes")

        u.groups.add(grupo_clientes)

        # Enviar correo de bienvenida al registrarse por este flujo simple
        try:
            if email:
                asunto = "¡Bienvenido a EpicAnimes!"
                cta_url = request.build_absolute_uri(reverse("login"))
                html_body = render_to_string(
                    "emails/signup_welcome.html",
                    {"username": username, "cta_url": cta_url, "support_email": settings.EMAIL_HOST_USER},
                )
                text_body = strip_tags(html_body)
                msg = EmailMultiAlternatives(asunto, text_body, settings.DEFAULT_FROM_EMAIL, [email])
                msg.attach_alternative(html_body, "text/html")
                msg.send(fail_silently=True)
        except Exception:
            logger.exception("No se pudo enviar correo de bienvenida (signup simple) a %s", email)

        messages.success(request, "Cuenta creada. Ahora puedes iniciar sesión.")

        return redirect("login")



    return render(request, "registration/signup.html")





def _procesar_compra(request, referencia_pago=None, datos_cliente=None, *, force=False):

    """Genera la orden y reduce stock al finalizar la compra."""
    if not force and obtener_rol_usuario(request.user) != "comprador":

        raise CarritoError("Solo los clientes pueden comprar.")

    cart = _get_cart(request)

    datos_normalizados = _resolver_datos_cliente(request, datos_cliente)

    request.session["checkout_info_prefill"] = datos_normalizados.copy()

    request.session.modified = True



    with transaction.atomic():

        lineas, total = _calcular_lineas_y_total(cart, lock=True)

        try:
            total, order_total, _, moneda_paypal, order_currency = _calcular_totales_paypal(total)
        except PayPalError as exc:
            raise CarritoError(str(exc))
        paso_moneda = paypal_amount_step(moneda_paypal)

        if referencia_pago:

            if not force:

                try:

                    captura = paypal_capture_order(

                        referencia_pago,

                        expected_amount=total,

                        expected_currency=moneda_paypal,

                    )

                except PayPalError as exc:

                    raise CarritoError(str(exc))

                referencia_pago = captura.capture_id or captura.order_id or referencia_pago

            if Compra.objects.select_for_update().filter(referencia_pago=referencia_pago).exists():

                raise CarritoError("Esta orden de pago ya fue procesada.")



        for producto, cantidad, _ in lineas:

            Compra.objects.create(

                cliente=datos_normalizados["nombre"],

                usuario=request.user,

                nombre_completo=datos_normalizados["nombre"],

                correo_contacto=datos_normalizados["correo"],

                telefono_contacto=datos_normalizados["telefono"],

                direccion_envio=datos_normalizados["direccion"],

                ciudad_envio=datos_normalizados["ciudad"],

                notas_extra=datos_normalizados["notas"],

                producto=producto,

                valor_producto=producto.precio,

                cantidad=cantidad,

                referencia_pago=referencia_pago,

            )

            if producto.vendedor_id:

                Venta.objects.create(

                    vendedor=producto.vendedor,

                    producto=producto,

                    cantidad=cantidad,

                )

            producto.existencias -= cantidad

            producto.save(update_fields=["existencias"])



    _save_cart(request, {})

    if paso_moneda == Decimal("1"):

        request.session["ultimo_total"] = f"{total:.0f}"

    else:

        request.session["ultimo_total"] = f"{total:.2f}"

    request.session["ultimo_checkout_info"] = {

        "nombre": datos_normalizados["nombre"],

        "email": datos_normalizados["correo"],

        "telefono": datos_normalizados["telefono"],

        "direccion": datos_normalizados["direccion"],

        "ciudad": datos_normalizados["ciudad"],

        "notas": datos_normalizados["notas"],

    }

    request.session.modified = True

    _notificar_actores_compra(

        datos_normalizados,

        lineas,

        total,

        referencia_pago,

        paso_moneda,

    )

    return total


def _calcular_totales_paypal(total):
    """Normaliza los totales según la configuración de PayPal."""
    moneda_paypal = getattr(settings, "PAYPAL_CURRENCY", "CLP")
    order_currency = getattr(settings, "PAYPAL_ORDER_CURRENCY", moneda_paypal)
    conversion_rate, _ = get_paypal_conversion_rate()
    total_normalizado, total_paypal = normalize_paypal_totals(
        total,
        conversion_rate=conversion_rate,
        store_currency=moneda_paypal,
        order_currency=order_currency,
    )
    return total_normalizado, total_paypal, conversion_rate, moneda_paypal, order_currency





@login_required

def VistaCarrito(request):

    """Muestra el carrito del usuario junto con la información de checkout."""
    cart = _get_cart(request)

    items, total, carrito_sin_fallos = _build_cart_items(cart)

    rol_actual = obtener_rol_usuario(request.user)

    es_comprador = rol_actual == "comprador"

    paypal_configurado, paypal_error = paypal_is_configured()



    perfil_cliente = None

    if request.user.is_authenticated:

        perfil_cliente = getattr(request.user, "perfil_cliente", None)

        if perfil_cliente is None:

            perfil_cliente = PerfilCliente.objects.filter(user=request.user).first()



    checkout_prefill = request.session.get("checkout_info_prefill", {}).copy()

    if request.user.is_authenticated:

        if perfil_cliente:

            checkout_prefill.setdefault("nombre", perfil_cliente.nombre or request.user.get_full_name() or request.user.username)

            checkout_prefill.setdefault("email", perfil_cliente.email or request.user.email or "")

            checkout_prefill.setdefault("telefono", perfil_cliente.telefono or "")

            checkout_prefill.setdefault("direccion", perfil_cliente.direccion or "")

            checkout_prefill.setdefault("ciudad", perfil_cliente.ciudad or "")

        else:

            checkout_prefill.setdefault("nombre", request.user.get_full_name() or request.user.username)

            checkout_prefill.setdefault("email", request.user.email or "")

    checkout_prefill.setdefault("telefono", checkout_prefill.get("telefono", ""))

    checkout_prefill.setdefault("direccion", checkout_prefill.get("direccion", ""))

    checkout_prefill.setdefault("ciudad", checkout_prefill.get("ciudad", ""))

    checkout_prefill.setdefault("notas", checkout_prefill.get("notas", ""))



    puede_pagar = es_comprador and carrito_sin_fallos and bool(items)

    paypal_summary = paypal_conversion_summary(total)

    contexto = {

        "items": items,

        "total": total,

        "carrito_sin_fallos": carrito_sin_fallos,

        "paypal_client_id": getattr(settings, "PAYPAL_CLIENT_ID", ""),

        **paypal_summary,

        "paypal_enabled": paypal_configurado,

        "paypal_error": paypal_error,

        "cart_count": _cart_count(cart),

        "rol_usuario": rol_actual,

        "puede_comprar": es_comprador,
        "puede_pagar": puede_pagar,
        "puede_pagar_paypal": puede_pagar and paypal_configurado,
        "checkout_prefill": checkout_prefill,

    }

    return render(request, "public/carrito.html", contexto)





@login_required

@require_http_methods(["POST"])

def agregar_al_carrito(request, producto_id):

    """Agrega un producto al carrito del usuario."""
    if obtener_rol_usuario(request.user) != "comprador":

        messages.error(request, "Tu rol no permite comprar en la tienda.")

        return redirect("index")

    producto = get_object_or_404(Producto, pk=producto_id)

    try:

        cantidad = int(request.POST.get("cantidad", 1))

    except (TypeError, ValueError):

        cantidad = 1

    cantidad = max(1, cantidad)

    cart = _get_cart(request)

    actual = cart.get(str(producto.id), 0)

    if actual + cantidad > producto.existencias:

        messages.error(request, "No hay stock suficiente del producto seleccionado.", extra_tags="critico")

        return redirect("index")

    cart[str(producto.id)] = actual + cantidad

    _save_cart(request, cart)

    messages.success(request, f"{producto.nombre} agregado al carrito.")

    return redirect("carrito")





@login_required

@require_http_methods(["POST"])

def actualizar_carrito(request, producto_id):

    """Actualiza las cantidades de un producto dentro del carrito."""
    is_json = request.headers.get("x-requested-with") == "XMLHttpRequest" or (request.content_type or "").startswith("application/json")

    if obtener_rol_usuario(request.user) != "comprador":

        if is_json:

            return JsonResponse({"ok": False, "error": "Tu rol no permite modificar el carrito."}, status=400)

        messages.error(request, "Tu rol no permite modificar el carrito.")

        return redirect("carrito")

    cart = _get_cart(request)

    key = str(producto_id)

    if key not in cart:

        if is_json:

            return JsonResponse({"ok": False, "error": "El producto no esta en tu carrito."}, status=404)

        messages.error(request, "El producto no esta en tu carrito.")

        return redirect("carrito")



    if is_json:

        try:

            payload = json.loads(request.body or "{}")

        except json.JSONDecodeError:

            payload = {}

        cantidad_raw = payload.get("cantidad", cart[key])

    else:

        cantidad_raw = request.POST.get("cantidad", cart[key])

    try:

        cantidad = int(cantidad_raw)

    except (TypeError, ValueError):

        cantidad = cart[key]



    if cantidad <= 0:

        cart.pop(key, None)

    else:

        producto = get_object_or_404(Producto, pk=producto_id)

        if cantidad > producto.existencias:

            if is_json:

                return JsonResponse({"ok": False, "error": "Stock insuficiente para el producto."}, status=400)

            messages.error(request, "Stock insuficiente para el producto.")

            return redirect("carrito")

        cart[key] = cantidad

    _save_cart(request, cart)



    if is_json:

        items, total, puede_pagar = _build_cart_items(cart)

        subtotal = Decimal("0")

        sin_stock = False

        stock_disponible = None

        for item in items:

            if item["producto"].id == int(producto_id):

                subtotal = item["subtotal"]

                sin_stock = item["sin_stock"]

                stock_disponible = item["producto"].existencias

                break

        return JsonResponse({

            "ok": True,

            "cantidad": cart.get(key, 0),

            "subtotal": f"{subtotal:.0f}",

            "subtotal_raw": f"{subtotal:.2f}",

            "total": f"{total:.0f}",

            "total_raw": f"{total:.2f}",

            "sin_stock": sin_stock,

            "stock_disponible": stock_disponible,

            "puede_pagar": puede_pagar,

            "cart_count": _cart_count(cart),

        })



    messages.success(request, "Carrito actualizado.")

    return redirect("carrito")

@login_required

@require_http_methods(["POST"])

def eliminar_del_carrito(request, producto_id):

    """Elimina un producto del carrito."""
    is_json = request.headers.get("x-requested-with") == "XMLHttpRequest" or (request.content_type or "").startswith("application/json")

    cart = _get_cart(request)

    key = str(producto_id)

    if key in cart:

        cart.pop(key)

        _save_cart(request, cart)

        if is_json:

            items, total, puede_pagar = _build_cart_items(cart)

            return JsonResponse({

                "ok": True,

                "total_raw": f"{total:.2f}",

                "cart_count": _cart_count(cart),

                "puede_pagar": puede_pagar,

                "items_restantes": len(items),

            })

        messages.success(request, "Producto eliminado del carrito.")

    else:

        if is_json:

            return JsonResponse({"ok": False, "error": "El producto ya no estaba en tu carrito."}, status=404)

        messages.info(request, "El producto ya no estaba en tu carrito.")

    return redirect("carrito")





@login_required

@require_http_methods(["POST"])

def paypal_crear_orden(request):

    """Crea una orden de PayPal para el contenido del carrito."""
    if obtener_rol_usuario(request.user) != "comprador":

        return JsonResponse({"ok": False, "error": "Tu rol no permite comprar en la tienda."}, status=403)



    content_type = request.content_type or ""

    cart = _get_cart(request)

    print("[paypal_crear_orden] usuario=", getattr(request.user, "id", None), "cart_keys=", list(cart.keys()))

    if content_type.startswith("application/json"):

        try:

            payload = json.loads(request.body or "{}")

        except json.JSONDecodeError:

            payload = {}

        datos_cliente = payload.get("datos_cliente") or {}

    else:

        datos_cliente = {

            "nombre": request.POST.get("nombre") or "",

            "email": request.POST.get("email") or "",

            "telefono": request.POST.get("telefono") or "",

            "direccion": request.POST.get("direccion") or "",

            "ciudad": request.POST.get("ciudad") or "",

            "notas": request.POST.get("notas") or "",

        }



    try:

        datos_normalizados = _resolver_datos_cliente(request, datos_cliente)

        _lineas, total = _calcular_lineas_y_total(cart, lock=False)

    except CarritoError as exc:

        print("[paypal_crear_orden] error datos/carrito:", exc)

        return JsonResponse({"ok": False, "error": str(exc)}, status=400)



    try:
        total, order_total, conversion_rate, moneda_paypal, order_currency = _calcular_totales_paypal(total)
    except PayPalError as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)

    print("[paypal_crear_orden] total_normalizado=", total, "moneda=", moneda_paypal)



    shipping = {

        "name": {"full_name": datos_normalizados["nombre"][:300]},

        "address": {

            "address_line_1": datos_normalizados["direccion"][:300] or "Direccion pendiente",

            "admin_area_1": "RM",

            "admin_area_2": datos_normalizados["ciudad"][:120] or "Santiago",

            "postal_code": "8320000",

            "country_code": "CL",

        },

    }

    try:

        reference = f"ORD-{request.user.id}-{timezone.now().strftime('%Y%m%d%H%M%S')}"

        print(

            "[paypal_crear_orden] order_total=",

            order_total,

            "order_currency=",

            order_currency,

            "conversion_rate=",

            conversion_rate,

        )

        order_id = paypal_create_order(order_total, order_currency, shipping=shipping, reference=reference)

    except PayPalError as exc:

        print("[paypal_crear_orden] error paypal:", exc)

        return JsonResponse({"ok": False, "error": str(exc)}, status=400)



    request.session["checkout_info_prefill"] = datos_normalizados.copy()

    request.session.modified = True



    return JsonResponse({"ok": True, "orderID": order_id})





@login_required

@require_http_methods(["POST"])

def finalizar_compra(request):

    """Valida el carrito y crea la orden de compra definitiva."""
    content_type = request.content_type or ""

    is_json = content_type.startswith("application/json") or request.headers.get("x-requested-with") == "XMLHttpRequest"

    referencia = None

    datos_cliente = {}

    if content_type.startswith("application/json"):

        try:

            payload = json.loads(request.body or "{}")

        except json.JSONDecodeError:

            payload = {}

        referencia = payload.get("paypal_order_id") or payload.get("orderID")

        datos_cliente = payload.get("datos_cliente") or {}

    else:

        referencia = request.POST.get("paypal_order_id")

        datos_cliente = {

            "nombre": request.POST.get("nombre") or "",

            "email": request.POST.get("email") or "",

            "telefono": request.POST.get("telefono") or "",

            "direccion": request.POST.get("direccion") or "",

            "ciudad": request.POST.get("ciudad") or "",

            "notas": request.POST.get("notas") or "",

        }

    if not referencia:

        mensaje_error = "No se recibio la confirmacion de PayPal. Intenta nuevamente."

        if is_json:

            return JsonResponse({"ok": False, "error": mensaje_error}, status=400)

        messages.error(request, mensaje_error)

        return redirect("carrito")



    try:

        total = _procesar_compra(request, referencia_pago=referencia, datos_cliente=datos_cliente)

    except CarritoError as exc:

        if is_json:

            return JsonResponse({"ok": False, "error": str(exc)}, status=400)

        messages.error(request, str(exc))

        return redirect("carrito")

    except Exception:

        if is_json:

            return JsonResponse({"ok": False, "error": "Ocurrio un error al procesar la compra."}, status=500)

        messages.error(request, "Ocurrio un error al procesar la compra.")

        return redirect("carrito")



    if is_json:

        return JsonResponse({"ok": True, "redirect": reverse("carrito_gracias"), "total": f"{total:.2f}"})



    messages.success(request, "Compra realizada correctamente.")

    return redirect("carrito_gracias")


@login_required
@require_http_methods(["POST"])
def carrito_compra_ficticia(request):
    """Permite registrar una compra de prueba sin pasar por PayPal."""
    content_type = request.content_type or ""
    is_json = content_type.startswith("application/json") or request.headers.get("x-requested-with") == "XMLHttpRequest"

    datos_cliente = {}
    if content_type.startswith("application/json"):
        try:
            payload = json.loads(request.body or "{}")
        except json.JSONDecodeError:
            payload = {}
        datos_cliente = payload.get("datos_cliente") or {}
    else:
        datos_cliente = {
            "nombre": request.POST.get("nombre") or "",
            "email": request.POST.get("email") or "",
            "telefono": request.POST.get("telefono") or "",
            "direccion": request.POST.get("direccion") or "",
            "ciudad": request.POST.get("ciudad") or "",
            "notas": request.POST.get("notas") or "",
        }

    referencia = f"FICT-{request.user.id}-{timezone.now().strftime('%Y%m%d%H%M%S')}"
    try:
        total = _procesar_compra(
            request,
            referencia_pago=referencia,
            datos_cliente=datos_cliente,
            force=True,
        )
    except CarritoError as exc:
        if is_json:
            return JsonResponse({"ok": False, "error": str(exc)}, status=400)
        messages.error(request, str(exc))
        return redirect("carrito")
    except Exception:
        logger.exception("Error al registrar compra ficticia")
        if is_json:
            return JsonResponse({"ok": False, "error": "No se pudo registrar la compra ficticia."}, status=500)
        messages.error(request, "No se pudo registrar la compra ficticia.")
        return redirect("carrito")

    mensaje_ok = "Compra ficticia registrada. Stock actualizado."
    if is_json:
        return JsonResponse(
            {"ok": True, "redirect": reverse("carrito_gracias"), "total": f"{Decimal(total):.2f}"},
            status=201,
        )
    messages.success(request, mensaje_ok)
    return redirect("carrito_gracias")


def carrito_gracias(request):

    """Muestra la página de confirmación después de pagar."""
    total = request.session.pop("ultimo_total", None)
    checkout_info = request.session.pop("ultimo_checkout_info", None)
    cart = _get_cart(request)
    contexto = {
        "total": total,
        "checkout_info": checkout_info,
        "cart_count": _cart_count(cart),
        "rol_usuario": obtener_rol_usuario(request.user),
    }
    return render(request, "public/carrito_gracias.html", contexto)


                                                              

                                      

                                                              

def VistaSobreNosotros(request):
    """Expone la página informativa de la empresa."""
    cart = _get_cart(request)
    contexto = {
        "cart_count": _cart_count(cart),
        "rol_usuario": obtener_rol_usuario(request.user),
        "nav_page": "sobre_nosotros",
    }
    return render(request, "public/sobrenosotros.html", contexto)


def VistaTerminos(request):
    """Entrega la vista con términos y condiciones."""
    cart = _get_cart(request)
    contexto = {
        "cart_count": _cart_count(cart),
        "rol_usuario": obtener_rol_usuario(request.user),
        "nav_page": "terminos",
    }
    return render(request, "public/terminos.html", contexto)





@require_http_methods(["POST"])
def send_login_otp(request):
    """Envía un código de verificación (6 dígitos) al correo del usuario indicado."""

    ident = (request.POST.get("username") or request.POST.get("email") or "").strip()

    if not ident:
        return JsonResponse({"error": "Falta usuario o email"}, status=400)

    try:
        try:
            user = User.objects.get(username=ident)
        except User.DoesNotExist:
            user = User.objects.get(email__iexact=ident)
    except User.DoesNotExist:
        return JsonResponse({"error": "Usuario no encontrado"}, status=404)

    if not user.email:
        return JsonResponse({"error": "Este usuario no tiene email registrado"}, status=400)

    import random

    code = random.randint(100000, 999999)
    expires_at = timezone.now() + timedelta(minutes=5)
    cache.set(
        f"login_otp:{user.id}",
        {"code": str(code), "expires_at": expires_at},
        300,
    )

    asunto = "Código de verificación EpicAnimes"
    cuerpo = (
        f"Hola {user.username},\n\n"
        f"Tu código de verificación es: {code}.\n"
        "Este código solo funciona 1 vez. Si no solicitaste este código, ignora este mensaje.\n\n"
        "EpicAnimes"
    )

    from_email = getattr(settings, "EMAIL_HOST_USER", None)
    display_from = getattr(settings, "DEFAULT_FROM_EMAIL", None)
    if from_email:
        display_from = display_from or f"EpicAnimes <{from_email}>"

    try:
        send_mail(
            asunto,
            cuerpo,
            display_from or from_email or "no-reply@epicanimes.local",
            [user.email],
            fail_silently=False,
        )
    except Exception as exc:
        logger.exception("Error enviando OTP a %s: %s", user.username, exc)
        return JsonResponse({"error": "No se pudo enviar el correo. Intenta más tarde."}, status=500)

    return JsonResponse({"ok": True})





@require_http_methods(["GET", "POST"])

def VistaContacto(request):

    """Renderiza la página de contacto con datos de soporte."""
    cart = _get_cart(request)

    cart_count = _cart_count(cart)

    rol_usuario = obtener_rol_usuario(request.user)

    form = PostulacionVendedorForm(request.POST or None)

    if request.method == "POST":

        if form.is_valid():

            form.save()

            messages.success(request, "Gracias por postular. Te contactaremos pronto.")

            return redirect("contacto")

        else:

            messages.error(request, "Revisa los datos del formulario.")

    contexto = {

        "form": form,

        "cart_count": cart_count,

        "rol_usuario": rol_usuario,

    }

    return render(request, "public/contacto.html", contexto)





def VistaProductoDetalle(request, producto_id):

    """Presenta el detalle de un producto específico."""
    producto = get_object_or_404(

        Producto.objects.select_related("vendedor", "vendedor__usuario"),

        pk=producto_id,

    )

    rol_usuario = obtener_rol_usuario(request.user)

    cart = _get_cart(request)

    cart_count = _cart_count(cart)

    default_image_url = static("images/Imagen1.png")



    producto.imagen_url = producto.imagen.url if getattr(producto, "imagen", None) else default_image_url

    vendedor_nombre = None

    if producto.vendedor and producto.vendedor.usuario:

        vendedor_nombre = producto.vendedor.usuario.get_full_name() or producto.vendedor.usuario.username



    relacionados_qs = (

        Producto.objects.exclude(pk=producto.pk)

        .filter(categoria__iexact=producto.categoria)

        .order_by("-fecha_ingreso")[:4]

    )

    relacionados = list(relacionados_qs)

    for rel in relacionados:

        rel.imagen_url = rel.imagen.url if getattr(rel, "imagen", None) else default_image_url



    contexto = {

        "producto": producto,

        "vendedor_nombre": vendedor_nombre,

        "relacionados": relacionados,

        "rol_usuario": rol_usuario,

        "puede_comprar": rol_usuario == "comprador",

        "cart_count": cart_count,

    }

    return render(request, "public/producto_detalle.html", contexto)





@require_http_methods(["GET", "POST"])

def VistaRegistro(request):

    """

    Registro simple de usuarios (self-service).

    Crea un usuario activo y redirige al login.

    """

    if request.method == "POST":

        username = (request.POST.get("username") or "").strip()
        email = (request.POST.get("email") or "").strip()
        p1 = (request.POST.get("password1") or "").strip()
        p2 = (request.POST.get("password2") or "").strip()



        errores = []

        if not username:
            errores.append("El nombre de usuario es obligatorio.")

        if not email:
            errores.append("El correo electrónico es obligatorio.")
        else:
            try:
                validate_email(email)
            except ValidationError:
                errores.append("Ingresa un correo electrónico válido.")

        if email and User.objects.filter(email__iexact=email).exists():
            errores.append("El correo electrónico ya está registrado.")

        if not p1:
            errores.append("La contraseña es obligatoria.")

        if p1 and len(p1) < 6:
            errores.append("La contraseña debe tener al menos 6 caracteres.")

        if p1 != p2:
            errores.append("Las contraseñas no coinciden.")

        if User.objects.filter(username=username).exists():
            errores.append("El nombre de usuario ya existe.")



        if errores:

            for e in errores:

                messages.error(request, e)

                                            

            return render(request, "registration/signup.html", {

                "form": {"username": username, "email": email}

            })



        u = User.objects.create_user(username=username, email=email, password=p1)

        u.is_active = True

        u.save()

        grupo_clientes, _ = Group.objects.get_or_create(name="Clientes")

        u.groups.add(grupo_clientes)

        messages.success(request, "Cuenta creada. Ahora puedes iniciar sesión.")

        return redirect("login")



    return render(request, "registration/signup.html")





def _procesar_compra(request, referencia_pago=None, datos_cliente=None, *, force=False):

    """Genera la orden y reduce stock al finalizar la compra."""
    if not force and obtener_rol_usuario(request.user) != "comprador":

        raise CarritoError("Solo los clientes pueden comprar.")

    cart = _get_cart(request)

    datos_normalizados = _resolver_datos_cliente(request, datos_cliente)

    request.session["checkout_info_prefill"] = datos_normalizados.copy()

    request.session.modified = True



    with transaction.atomic():

        lineas, total = _calcular_lineas_y_total(cart, lock=True)

        try:
            total, order_total, _, moneda_paypal, order_currency = _calcular_totales_paypal(total)
        except PayPalError as exc:
            raise CarritoError(str(exc))
        paso_moneda = paypal_amount_step(moneda_paypal)

        if referencia_pago:

            if not force:

                try:

                    captura = paypal_capture_order(

                        referencia_pago,

                        expected_amount=total,

                        expected_currency=moneda_paypal,

                    )

                except PayPalError as exc:

                    raise CarritoError(str(exc))

                referencia_pago = captura.capture_id or captura.order_id or referencia_pago

            if Compra.objects.select_for_update().filter(referencia_pago=referencia_pago).exists():

                raise CarritoError("Esta orden de pago ya fue procesada.")



        for producto, cantidad, _ in lineas:

            Compra.objects.create(

                cliente=datos_normalizados["nombre"],

                usuario=request.user,

                nombre_completo=datos_normalizados["nombre"],

                correo_contacto=datos_normalizados["correo"],

                telefono_contacto=datos_normalizados["telefono"],

                direccion_envio=datos_normalizados["direccion"],

                ciudad_envio=datos_normalizados["ciudad"],

                notas_extra=datos_normalizados["notas"],

                producto=producto,

                valor_producto=producto.precio,

                cantidad=cantidad,

                referencia_pago=referencia_pago,

            )

            if producto.vendedor_id:

                Venta.objects.create(

                    vendedor=producto.vendedor,

                    producto=producto,

                    cantidad=cantidad,

                )

            producto.existencias -= cantidad

            producto.save(update_fields=["existencias"])



    _save_cart(request, {})

    if paso_moneda == Decimal("1"):

        request.session["ultimo_total"] = f"{total:.0f}"

    else:

        request.session["ultimo_total"] = f"{total:.2f}"

    request.session["ultimo_checkout_info"] = {

        "nombre": datos_normalizados["nombre"],

        "email": datos_normalizados["correo"],

        "telefono": datos_normalizados["telefono"],

        "direccion": datos_normalizados["direccion"],

        "ciudad": datos_normalizados["ciudad"],

        "notas": datos_normalizados["notas"],

    }

    request.session.modified = True

    _notificar_actores_compra(

        datos_normalizados,

        lineas,

        total,

        referencia_pago,

        paso_moneda,

    )

    return total





@login_required

def VistaCarrito(request):

    """Muestra el carrito del usuario junto con la información de checkout."""
    cart = _get_cart(request)

    items, total, carrito_sin_fallos = _build_cart_items(cart)

    rol_actual = obtener_rol_usuario(request.user)

    es_comprador = rol_actual == "comprador"

    paypal_configurado, paypal_error = paypal_is_configured()



    perfil_cliente = None

    if request.user.is_authenticated:

        perfil_cliente = getattr(request.user, "perfil_cliente", None)

        if perfil_cliente is None:

            perfil_cliente = PerfilCliente.objects.filter(user=request.user).first()



    checkout_prefill = request.session.get("checkout_info_prefill", {}).copy()

    if request.user.is_authenticated:

        if perfil_cliente:

            checkout_prefill.setdefault("nombre", perfil_cliente.nombre or request.user.get_full_name() or request.user.username)

            checkout_prefill.setdefault("email", perfil_cliente.email or request.user.email or "")

            checkout_prefill.setdefault("telefono", perfil_cliente.telefono or "")

            checkout_prefill.setdefault("direccion", perfil_cliente.direccion or "")

            checkout_prefill.setdefault("ciudad", perfil_cliente.ciudad or "")

        else:

            checkout_prefill.setdefault("nombre", request.user.get_full_name() or request.user.username)

            checkout_prefill.setdefault("email", request.user.email or "")

    checkout_prefill.setdefault("telefono", checkout_prefill.get("telefono", ""))

    checkout_prefill.setdefault("direccion", checkout_prefill.get("direccion", ""))

    checkout_prefill.setdefault("ciudad", checkout_prefill.get("ciudad", ""))

    checkout_prefill.setdefault("notas", checkout_prefill.get("notas", ""))



    puede_pagar = es_comprador and carrito_sin_fallos and bool(items)

    paypal_summary = paypal_conversion_summary(total)

    contexto = {

        "items": items,

        "total": total,

        "carrito_sin_fallos": carrito_sin_fallos,

        "paypal_client_id": getattr(settings, "PAYPAL_CLIENT_ID", ""),

        **paypal_summary,
        "paypal_enabled": paypal_configurado,

        "paypal_error": paypal_error,

        "cart_count": _cart_count(cart),

        "rol_usuario": rol_actual,

        "puede_comprar": es_comprador,

        "puede_pagar": puede_pagar,
        "puede_pagar_paypal": puede_pagar and paypal_configurado,
        "checkout_prefill": checkout_prefill,

    }

    return render(request, "public/carrito.html", contexto)





@login_required

@require_http_methods(["POST"])

def agregar_al_carrito(request, producto_id):

    """Agrega un producto al carrito del usuario."""
    if obtener_rol_usuario(request.user) != "comprador":

        messages.error(request, "Tu rol no permite comprar en la tienda.")

        return redirect("index")

    producto = get_object_or_404(Producto, pk=producto_id)

    try:

        cantidad = int(request.POST.get("cantidad", 1))

    except (TypeError, ValueError):

        cantidad = 1

    cantidad = max(1, cantidad)

    cart = _get_cart(request)

    actual = cart.get(str(producto.id), 0)

    if actual + cantidad > producto.existencias:

        messages.error(request, "No hay stock suficiente del producto seleccionado.", extra_tags="critico")

        return redirect("index")

    cart[str(producto.id)] = actual + cantidad

    _save_cart(request, cart)

    messages.success(request, f"{producto.nombre} agregado al carrito.")

    return redirect("carrito")





@login_required

@require_http_methods(["POST"])

def actualizar_carrito(request, producto_id):

    """Actualiza las cantidades de un producto dentro del carrito."""
    is_json = request.headers.get("x-requested-with") == "XMLHttpRequest" or (request.content_type or "").startswith("application/json")

    if obtener_rol_usuario(request.user) != "comprador":

        if is_json:

            return JsonResponse({"ok": False, "error": "Tu rol no permite modificar el carrito."}, status=400)

        messages.error(request, "Tu rol no permite modificar el carrito.")

        return redirect("carrito")

    cart = _get_cart(request)

    key = str(producto_id)

    if key not in cart:

        if is_json:

            return JsonResponse({"ok": False, "error": "El producto no esta en tu carrito."}, status=404)

        messages.error(request, "El producto no esta en tu carrito.")

        return redirect("carrito")



    if is_json:

        try:

            payload = json.loads(request.body or "{}")

        except json.JSONDecodeError:

            payload = {}

        cantidad_raw = payload.get("cantidad", cart[key])

    else:

        cantidad_raw = request.POST.get("cantidad", cart[key])

    try:

        cantidad = int(cantidad_raw)

    except (TypeError, ValueError):

        cantidad = cart[key]



    if cantidad <= 0:

        cart.pop(key, None)

    else:

        producto = get_object_or_404(Producto, pk=producto_id)

        if cantidad > producto.existencias:

            if is_json:

                return JsonResponse({"ok": False, "error": "Stock insuficiente para el producto."}, status=400)

            messages.error(request, "Stock insuficiente para el producto.")

            return redirect("carrito")

        cart[key] = cantidad

    _save_cart(request, cart)



    if is_json:

        items, total, puede_pagar = _build_cart_items(cart)

        subtotal = Decimal("0")

        sin_stock = False

        stock_disponible = None

        for item in items:

            if item["producto"].id == int(producto_id):

                subtotal = item["subtotal"]

                sin_stock = item["sin_stock"]

                stock_disponible = item["producto"].existencias

                break

        return JsonResponse({

            "ok": True,

            "cantidad": cart.get(key, 0),

            "subtotal": f"{subtotal:.0f}",

            "subtotal_raw": f"{subtotal:.2f}",

            "total": f"{total:.0f}",

            "total_raw": f"{total:.2f}",

            "sin_stock": sin_stock,

            "stock_disponible": stock_disponible,

            "puede_pagar": puede_pagar,

            "cart_count": _cart_count(cart),

        })



    messages.success(request, "Carrito actualizado.")

    return redirect("carrito")

@login_required

@require_http_methods(["POST"])

def eliminar_del_carrito(request, producto_id):

    """Elimina un producto del carrito."""
    is_json = request.headers.get("x-requested-with") == "XMLHttpRequest" or (request.content_type or "").startswith("application/json")

    cart = _get_cart(request)

    key = str(producto_id)

    if key in cart:

        cart.pop(key)

        _save_cart(request, cart)

        if is_json:

            items, total, puede_pagar = _build_cart_items(cart)

            return JsonResponse({

                "ok": True,

                "total_raw": f"{total:.2f}",

                "cart_count": _cart_count(cart),

                "puede_pagar": puede_pagar,

                "items_restantes": len(items),

            })

        messages.success(request, "Producto eliminado del carrito.")

    else:

        if is_json:

            return JsonResponse({"ok": False, "error": "El producto ya no estaba en tu carrito."}, status=404)

        messages.info(request, "El producto ya no estaba en tu carrito.")

    return redirect("carrito")





@login_required

@require_http_methods(["POST"])

def paypal_crear_orden(request):

    """Crea una orden de PayPal para el contenido del carrito."""
    if obtener_rol_usuario(request.user) != "comprador":

        return JsonResponse({"ok": False, "error": "Tu rol no permite comprar en la tienda."}, status=403)



    content_type = request.content_type or ""

    cart = _get_cart(request)

    print("[paypal_crear_orden] usuario=", getattr(request.user, "id", None), "cart_keys=", list(cart.keys()))

    if content_type.startswith("application/json"):

        try:

            payload = json.loads(request.body or "{}")

        except json.JSONDecodeError:

            payload = {}

        datos_cliente = payload.get("datos_cliente") or {}

    else:

        datos_cliente = {

            "nombre": request.POST.get("nombre") or "",

            "email": request.POST.get("email") or "",

            "telefono": request.POST.get("telefono") or "",

            "direccion": request.POST.get("direccion") or "",

            "ciudad": request.POST.get("ciudad") or "",

            "notas": request.POST.get("notas") or "",

        }



    try:

        datos_normalizados = _resolver_datos_cliente(request, datos_cliente)

        _lineas, total = _calcular_lineas_y_total(cart, lock=False)

    except CarritoError as exc:

        print("[paypal_crear_orden] error datos/carrito:", exc)

        return JsonResponse({"ok": False, "error": str(exc)}, status=400)



    moneda_paypal = getattr(settings, "PAYPAL_CURRENCY", "CLP")

    try:
        total, order_total, conversion_rate, moneda_paypal, order_currency = _calcular_totales_paypal(total)
    except PayPalError as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)

    print("[paypal_crear_orden] total_normalizado=", total, "moneda=", moneda_paypal)



    shipping = {

        "name": {"full_name": datos_normalizados["nombre"][:300]},

        "address": {

            "address_line_1": datos_normalizados["direccion"][:300] or "Direccion pendiente",

            "admin_area_1": "RM",

            "admin_area_2": datos_normalizados["ciudad"][:120] or "Santiago",

            "postal_code": "8320000",

            "country_code": "CL",

        },

    }

    try:

        reference = f"ORD-{request.user.id}-{timezone.now().strftime('%Y%m%d%H%M%S')}"

        print(

            "[paypal_crear_orden] order_total=",

            order_total,

            "order_currency=",

            order_currency,

            "conversion_rate=",

            conversion_rate,

        )

        order_id = paypal_create_order(order_total, order_currency, shipping=shipping, reference=reference)

    except PayPalError as exc:

        print("[paypal_crear_orden] error paypal:", exc)

        return JsonResponse({"ok": False, "error": str(exc)}, status=400)



    request.session["checkout_info_prefill"] = datos_normalizados.copy()

    request.session.modified = True



    return JsonResponse({"ok": True, "orderID": order_id})





@login_required

@require_http_methods(["POST"])

def finalizar_compra(request):

    """Valida el carrito y crea la orden de compra definitiva."""
    content_type = request.content_type or ""

    is_json = content_type.startswith("application/json") or request.headers.get("x-requested-with") == "XMLHttpRequest"

    referencia = None

    datos_cliente = {}

    if content_type.startswith("application/json"):

        try:

            payload = json.loads(request.body or "{}")

        except json.JSONDecodeError:

            payload = {}

        referencia = payload.get("paypal_order_id") or payload.get("orderID")

        datos_cliente = payload.get("datos_cliente") or {}

    else:

        referencia = request.POST.get("paypal_order_id")

        datos_cliente = {

            "nombre": request.POST.get("nombre") or "",

            "email": request.POST.get("email") or "",

            "telefono": request.POST.get("telefono") or "",

            "direccion": request.POST.get("direccion") or "",

            "ciudad": request.POST.get("ciudad") or "",

            "notas": request.POST.get("notas") or "",

        }

    if not referencia:

        mensaje_error = "No se recibio la confirmacion de PayPal. Intenta nuevamente."

        if is_json:

            return JsonResponse({"ok": False, "error": mensaje_error}, status=400)

        messages.error(request, mensaje_error)

        return redirect("carrito")



    try:

        total = _procesar_compra(request, referencia_pago=referencia, datos_cliente=datos_cliente)

    except CarritoError as exc:

        if is_json:

            return JsonResponse({"ok": False, "error": str(exc)}, status=400)

        messages.error(request, str(exc))

        return redirect("carrito")

    except Exception:

        if is_json:

            return JsonResponse({"ok": False, "error": "Ocurrio un error al procesar la compra."}, status=500)

        messages.error(request, "Ocurrio un error al procesar la compra.")

        return redirect("carrito")



    if is_json:

        return JsonResponse({"ok": True, "redirect": reverse("carrito_gracias"), "total": f"{total:.2f}"})



    messages.success(request, "Compra realizada correctamente.")

    return redirect("carrito_gracias")





@login_required

def carrito_gracias(request):

    """Muestra la página de confirmación después de pagar."""
    total = request.session.pop("ultimo_total", None)

    checkout_info = request.session.pop("ultimo_checkout_info", None)

    return render(request, "public/carrito_gracias.html", {"total": total, "checkout_info": checkout_info})








                                                              

                                       

                                                   

                                                              



@login_required

def VistaVendedor(request):

    """Renderiza el panel privado para los vendedores."""
    vendedor = Vendedor.objects.filter(usuario=request.user).first()

    productos = Producto.objects.none()

    if vendedor:

        productos = Producto.objects.filter(vendedor=vendedor)

    form_data = {}



    if request.method == "POST":

        form_data = request.POST.dict()



        nombre = form_data.get("nombre", "").strip()

        descripcion = form_data.get("descripcion", "").strip()

        marca = form_data.get("marca", "").strip()

        fecha_raw = form_data.get("fecha_ingreso", "").strip()

        calidad = form_data.get("calidad", "").strip()

        precio_raw = form_data.get("precio", "").strip()

        existencias_raw = form_data.get("existencias", "").strip()

        categoria = form_data.get("categoria", "").strip()

        imagen = request.FILES.get("imagen")



        errores = []
        if imagen:
            try:
                _validar_imagen_producto(imagen)
            except ValidationError as exc:
                errores.append(str(exc))

        if not nombre:

            errores.append("El nombre del producto es obligatorio.")

        if not marca:

            errores.append("La marca es obligatoria.")

        if not calidad:

            errores.append("Indica la calidad del producto.")

        if not categoria:

            errores.append("Selecciona una categoría.")



                          

        fecha_ingreso = None

        if fecha_raw:

            fecha_ingreso = parse_date(fecha_raw)

            if not fecha_ingreso:

                errores.append("La fecha de ingreso no tiene un formato válido (AAAA-MM-DD).")

        else:

            fecha_ingreso = timezone.localdate()



                

        precio = None

        if precio_raw:

            try:

                precio = Decimal(precio_raw)

                if precio < 0:

                    raise InvalidOperation

            except (InvalidOperation, TypeError):

                errores.append("Ingresa un precio válido.")

        else:

            errores.append("El precio es obligatorio.")



                     

        existencias = None

        if existencias_raw:

            try:

                existencias = int(existencias_raw)

                if existencias < 0:

                    raise ValueError

            except (TypeError, ValueError):

                errores.append("Las existencias deben ser un número entero positivo.")

        else:

            errores.append("Debes indicar existencias disponibles.")



        if errores:

            for error in errores:

                messages.error(request, error)

        else:

            if form_data.get("producto_id"):

                try:

                    p = Producto.objects.get(pk=form_data.get("producto_id"))

                except Producto.DoesNotExist:

                    messages.error(request, "Producto no encontrado.")

                    return redirect("dashboard_vendedor")

                                                

                if vendedor and p.vendedor_id and p.vendedor_id != vendedor.id:

                    messages.error(request, "No puedes editar productos de otro vendedor.")

                    return redirect("dashboard_vendedor")

                p.nombre = nombre

                p.descripcion = descripcion

                p.marca = marca

                p.fecha_ingreso = fecha_ingreso

                p.calidad = calidad

                p.precio = precio

                p.existencias = existencias

                p.categoria = categoria

                if vendedor and not p.vendedor_id:

                    p.vendedor = vendedor

                if imagen:

                    p.imagen = imagen

                p.save()

                messages.success(request, "Producto actualizado correctamente.")

                return redirect("dashboard_vendedor")

            else:

                Producto.objects.create(

                    vendedor=vendedor,

                    nombre=nombre,

                    descripcion=descripcion,

                    marca=marca,

                    fecha_ingreso=fecha_ingreso,

                    calidad=calidad,

                    precio=precio,

                    existencias=existencias,

                    categoria=categoria,

                    imagen=imagen,

                )

            messages.success(request, "Producto agregado correctamente.")

            return redirect("dashboard_vendedor")



    categorias_qs = Producto.objects.all()

    if vendedor:

        categorias_qs = categorias_qs.filter(vendedor=vendedor)

    categorias = categorias_qs.order_by("categoria").values_list("categoria", flat=True).distinct()



    contexto = {

        "productos": productos,

        "form_data": form_data,

        "categorias": categorias,

        "calidad_opciones": [

            "Nuevo",

            "Casi nuevo",

            "Coleccionista",

            "Usado",

        ],

        "today": timezone.localdate().isoformat(),

    }

    return render(request, "dashboards/dashboard_vendedor.html", contexto)





@login_required

@require_http_methods(["GET"])

def api_vendedor_resumen(request):

    """

    Devuelve métricas REALES para el vendedor actual:

    - ventas_hoy, ticket_promedio, serie úlltimos 7 días,

      ventas por categoría (top 5), totales.

    """

    vendedor = Vendedor.objects.filter(usuario=request.user).first()

    if not vendedor:

        return HttpResponseForbidden("No es vendedor")



    ventas_qs = Venta.objects.filter(vendedor=vendedor)



    total_ventas = float(ventas_qs.aggregate(s=Sum("total"))["s"] or 0)

    total_items = int(ventas_qs.aggregate(s=Sum("cantidad"))["s"] or 0)

    n_ventas = ventas_qs.count()

    ticket_prom = float((total_ventas / n_ventas) if n_ventas else 0)



                                        

    today = timezone.localdate()

    last7 = [today - timedelta(days=d) for d in range(6, -1, -1)]

    ventas_por_dia = ventas_qs.values("fecha_venta").annotate(total=Sum("total"))

    mapa_dias = {v["fecha_venta"]: float(v["total"] or 0) for v in ventas_por_dia}

    labels = [d.strftime("%a") for d in last7]                 

    data = [mapa_dias.get(d, 0) for d in last7]



                                

    por_categoria_qs = (

        ventas_qs

        .values(nombre=F("producto__categoria"))

        .annotate(total=Sum("total"))

        .order_by("-total")[:5]

    )

    por_categoria = [

        {"categoria": r["nombre"] or "Sin categoria", "total": float(r["total"] or 0)}

        for r in por_categoria_qs

    ]



    ventas_hoy = float(

        ventas_qs.filter(fecha_venta=today).aggregate(s=Sum("total"))["s"] or 0

    )



    return JsonResponse({

        "ventas_hoy": ventas_hoy,

        "ticket_promedio": ticket_prom,

        "tasa_conversion": 2.4,                                                      

        "labels": labels,

        "data": data,

        "por_categoria": por_categoria,

        "total_ventas": total_ventas,

        "total_items": total_items,

    })


                                                           
@login_required
@require_http_methods(["GET"])
def api_vendedor_resumen_ext(request):
    """Expone métricas extendidas del vendedor en JSON."""
    vendedor = Vendedor.objects.filter(usuario=request.user).first()
    if not vendedor:
        return HttpResponseForbidden("No es vendedor")

    ventas_qs = Venta.objects.filter(vendedor=vendedor)

    total_ventas = float(ventas_qs.aggregate(s=Sum("total"))["s"] or 0)
    total_items = int(ventas_qs.aggregate(s=Sum("cantidad"))["s"] or 0)
    n_ventas = ventas_qs.count()
    ticket_prom = float((total_ventas / n_ventas) if n_ventas else 0)

    today = timezone.localdate()
    try:
        days = int(request.GET.get("days", 7))
    except (TypeError, ValueError):
        days = 7
    days = max(7, min(days, 365))
    desde = today - timedelta(days=days - 1)

    ventas_por_dia = (
        ventas_qs
        .filter(fecha_venta__gte=desde, fecha_venta__lte=today)
        .values("fecha_venta")
        .annotate(total=Sum("total"))
    )
    dias = [desde + timedelta(days=i) for i in range(days)]
    mapa_dias = {v["fecha_venta"]: float(v["total"] or 0) for v in ventas_por_dia}
    labels = [d.strftime("%a") if days == 7 else d.strftime("%d/%m") for d in dias]
    data = [mapa_dias.get(d, 0.0) for d in dias]

    por_categoria_qs = (
        ventas_qs
        .filter(fecha_venta__gte=desde, fecha_venta__lte=today)
        .values(nombre=F("producto__categoria"))
        .annotate(total=Sum("total"))
        .order_by("-total")[:5]
    )
    por_categoria = [
        {"categoria": r["nombre"] or "Sin categoria", "total": float(r["total"] or 0)}
        for r in por_categoria_qs
    ]

    ventas_hoy = float(
        ventas_qs.filter(fecha_venta=today).aggregate(s=Sum("total"))["s"] or 0
    )

    return JsonResponse({
        "ventas_hoy": ventas_hoy,
        "ticket_promedio": ticket_prom,
        "tasa_conversion": 2.4,
        "labels": labels,
        "data": data,
        "por_categoria": por_categoria,
        "total_ventas": total_ventas,
        "total_items": total_items,
        "days": days,
    })




@login_required

@require_http_methods(["GET"])

def api_vendedor_producto_detalle(request, pk):

    """

    Detalle de un producto para prefijar el formulario de edición del vendedor.

    """

    vendedor = Vendedor.objects.filter(usuario=request.user).first()

    try:

        p = Producto.objects.get(pk=pk)

    except Producto.DoesNotExist:

        return JsonResponse({"error": "not_found"}, status=404)

    if vendedor and p.vendedor_id and p.vendedor_id != vendedor.id:

        return JsonResponse({"error": "forbidden"}, status=403)



    data = {

        "id": p.id,

        "nombre": p.nombre or "",

        "marca": p.marca or "",

        "calidad": p.calidad or "",

        "categoria": p.categoria or "",

        "precio": float(p.precio or 0),

        "existencias": int(p.existencias or 0),

        "fecha": (p.fecha_ingreso.isoformat() if getattr(p, "fecha_ingreso", None) else ""),

        "descripcion": p.descripcion or "",

    }

    return JsonResponse(data)





                                                              

                                           

                                    

                                                        

                                                              



@login_required

def VistaAdministrador(request):

    """

    Renderiza el panel admin (si lo usas render server-side).

    En general, los datos ahora se consumen vía las APIs de abajo.

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo administradores")

    total_vendedores = Vendedor.objects.count()

    total_productos = Producto.objects.count()

    total_existencias = Producto.objects.aggregate(s=Sum("existencias"))["s"] or 0

    total_ventas = Venta.objects.aggregate(s=Sum("total"))["s"] or 0

    total_items = Venta.objects.aggregate(s=Sum("cantidad"))["s"] or 0



                                           

    ventas_por_vendedor = (

        Venta.objects

        .values("vendedor__usuario__username")

        .annotate(total=Sum("total"), cantidad=Sum("cantidad"))

        .order_by("-total")

    )

    labels_vendedores = [v["vendedor__usuario__username"] or "N/D" for v in ventas_por_vendedor]

    data_vendedores = [float(v["total"] or 0) for v in ventas_por_vendedor]



                                             

    hoy = timezone.localdate()

    hace_30 = hoy - timedelta(days=30)

    clientes_totales = Compra.objects.values_list("cliente", flat=True).distinct().count()

    clientes_activos_30 = (

        Compra.objects

        .filter(fecha_compra__gte=hace_30)

        .values_list("cliente", flat=True).distinct().count()

    )

    clientes_inactivos_30 = max(clientes_totales - clientes_activos_30, 0)



                              

    productos_bajo_stock = Producto.objects.filter(existencias__lte=5).order_by("existencias", "nombre")



    contexto = {

        "total_vendedores": total_vendedores,

        "total_productos": total_productos,

        "total_existencias": total_existencias,

        "total_ventas": total_ventas,

        "total_items_vendidos": total_items,

        "labels_vendedores": labels_vendedores,

        "data_vendedores": data_vendedores,

        "tabla_vendedores": ventas_por_vendedor,

        "productos_bajo_stock": productos_bajo_stock,

        "clientes_activos_30": clientes_activos_30,

        "clientes_inactivos_30": clientes_inactivos_30,

    }

                                                                  

    contexto["usuarios_iniciales"] = []

    contexto["stock_inicial"] = []

    return render(request, "dashboards/dashboard_administrador.html", contexto)





                                                                 



def _json_body(request):

    """Intenta decodificar el cuerpo JSON de la solicitud."""
    try:

        return json.loads(request.body.decode("utf-8") or "{}")

    except Exception:

        return {}



def _bad_request(message: str):

    """Devuelve una respuesta JSON de error con estado 400."""
    try:

        msg = str(message)

    except Exception:

        msg = "bad_request"

                                              

    try:

        print(f"api_admin_vendedores 400 -> {msg}")

    except Exception:

        pass

    return JsonResponse({"error": msg}, status=400)





                                                                 



@login_required

@require_http_methods(["GET"])

def api_admin_productos_bajo_stock(request):

    """

    Listado de productos.

    - Por defecto: solo críticos (existencias <= 5).

    - ?all=1 → devuelve TODOS los productos.

    - ?vendedor_id=XX → filtra por productos que ese vendedor vendió alguna vez.

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



    vendedor_id = request.GET.get("vendedor_id")

    ver_todos = request.GET.get("all") in ("1", "true", "True")



    qs = Producto.objects.all() if ver_todos else Producto.objects.filter(existencias__lte=5)



    if vendedor_id:

        ventas_sub = Venta.objects.filter(producto_id=OuterRef("pk"), vendedor_id=vendedor_id)

        qs = qs.annotate(vendido_por=Exists(ventas_sub)).filter(vendido_por=True)



    qs = qs.select_related("vendedor__usuario").order_by("vendedor__usuario__username", "existencias", "nombre")



    items = []

    for p in qs:

        vendedor_nombre = getattr(getattr(p.vendedor, "usuario", None), "username", None)

        imagen_url = None

        try:

            if p.imagen and hasattr(p.imagen, "url"):

                imagen_url = p.imagen.url

        except Exception:

            imagen_url = None

        items.append({

            "id": p.id,

            "nombre": p.nombre,

            "vendedor": vendedor_nombre or "N/D",

            "stock": int(p.existencias or 0),

            "existencias": int(p.existencias or 0),

            "critico": bool((p.existencias or 0) <= 5),

            "tipo": p.categoria or "-",

            "categoria": p.categoria or "-",

            "imagen": imagen_url,

            "descripcion": p.descripcion or "",

        })

    return JsonResponse({"items": items})





@login_required

@require_http_methods(["GET"])

def api_admin_producto_detalle(request, pk):

    """Devuelve la información detallada de un producto."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")

    try:

        p = Producto.objects.get(pk=pk)

    except Producto.DoesNotExist:

        return JsonResponse({"error": "not_found"}, status=404)

    imagen_url = None

    try:

        if p.imagen and hasattr(p.imagen, "url"):

            imagen_url = p.imagen.url

    except Exception:

        imagen_url = None

    vendedor_nombre = getattr(getattr(p.vendedor, "usuario", None), "username", None)

    return JsonResponse({

        "id": p.id,

        "nombre": p.nombre or "",

        "existencias": int(p.existencias or 0),

        "critico": bool((p.existencias or 0) <= 5),

        "descripcion": p.descripcion or "",

        "imagen": imagen_url,

        "vendedor": vendedor_nombre or "N/D",

        "categoria": p.categoria or "",

    })





@login_required

@require_http_methods(["POST"])

def api_admin_producto_update_full(request, pk):

    """Actualiza todos los campos editables de un producto."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")

    try:

        p = Producto.objects.get(pk=pk)

    except Producto.DoesNotExist:

        return JsonResponse({"error": "not_found"}, status=404)



    nombre = request.POST.get("nombre", "").strip()

    descripcion = request.POST.get("descripcion", "").strip()

    existencias_raw = request.POST.get("existencias", "").strip()

    categoria = request.POST.get("categoria", "").strip()

    imagen = request.FILES.get("imagen")



    if nombre:

        p.nombre = nombre

    p.descripcion = descripcion

    if categoria:

        p.categoria = categoria

    if existencias_raw != "":

        try:

            p.existencias = int(existencias_raw)

            if p.existencias < 0:

                return HttpResponseBadRequest("existencias debe ser entero >= 0")

        except ValueError:

            return HttpResponseBadRequest("existencias inválidas")

    if imagen is not None:
        try:
            _validar_imagen_producto(imagen)
        except ValidationError as exc:
            return HttpResponseBadRequest(str(exc))
        p.imagen = imagen

    p.save()

    return JsonResponse({"ok": True, "id": p.id})





@login_required

@require_http_methods(["DELETE"])

def api_admin_producto_delete(request, pk):

    """Elimina un producto desde el panel administrativo."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")

    try:

        p = Producto.objects.get(pk=pk)

    except Producto.DoesNotExist:

        return JsonResponse({"error": "not_found"}, status=404)

    p.delete()

    return JsonResponse({"ok": True})





@login_required

@require_http_methods(["DELETE", "POST"])

def api_vendedor_producto_delete(request, pk):

    """Permite a un vendedor eliminar un producto propio.



    Acepta DELETE (preferido) y POST (fallback para entornos que no permiten DELETE desde formularios).

    """

    try:

        vend = Vendedor.objects.get(usuario=request.user)

    except Vendedor.DoesNotExist:

        return HttpResponseForbidden("Solo vendedores")



    try:

        p = Producto.objects.get(pk=pk, vendedor=vend)

    except Producto.DoesNotExist:

        return JsonResponse({"error": "not_found"}, status=404)



    p.delete()

    return JsonResponse({"ok": True})





@login_required

@require_http_methods(["PUT", "PATCH"])

def api_admin_producto_update_stock(request, pk):

    """

    Actualiza existencias de un producto (botón Editar en tabla de stock admin).

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



    data = _json_body(request)

    nuevo_stock = data.get("existencias")

    if nuevo_stock is None:

        return HttpResponseBadRequest("Campo 'existencias' requerido")



    try:

        nuevo_stock = int(nuevo_stock)

        if nuevo_stock < 0:

            raise ValueError

    except ValueError:

        return HttpResponseBadRequest("existencias debe ser entero >= 0")



    try:

        p = Producto.objects.get(pk=pk)

    except Producto.DoesNotExist:

        return JsonResponse({"error": "not_found"}, status=404)



    p.existencias = nuevo_stock

    p.save(update_fields=["existencias"])

    return JsonResponse({"ok": True, "id": p.id, "nombre": p.nombre, "existencias": p.existencias})





                                                                 



@login_required

@require_http_methods(["GET"])

def api_admin_ventas_por_vendedor(request):

    """

    Devuelve ventas por vendedor.

    - Solo vendedores ACTIVOS (User.is_active=True).

    - Incluye vendedores sin ventas (0).

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



    labels, data, rows = [], [], []

    vendedor_id = request.GET.get("vendedor_id")

    vendedores_qs = Vendedor.objects.select_related("usuario").filter(usuario__is_active=True)

    if vendedor_id:

        try:

            vendedores_qs = vendedores_qs.filter(pk=int(vendedor_id))

        except (TypeError, ValueError):

            vendedores_qs = vendedores_qs.none()

    vendedores_activos = vendedores_qs.order_by("usuario__username")

    for v in vendedores_activos:

        tot = Venta.objects.filter(vendedor=v).aggregate(s=Sum("total"), c=Sum("cantidad"))

        total = float(tot["s"] or 0)

        cant = int(tot["c"] or 0)

        nombre = v.usuario.username if v.usuario else f"Vendedor {v.id}"

        labels.append(nombre)

        data.append(total)

        rows.append({"vendedor": nombre, "total": total, "cantidad": cant})



    return JsonResponse({"labels": labels, "data": data, "rows": rows})





@login_required

@require_http_methods(["GET"])

@login_required

@require_http_methods(["GET"])

def api_admin_clientes_actividad(request):

    """

    Clientes activos vs inactivos últimos 30 días (si lo sigues mostrando en el donut).

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



    hoy = timezone.localdate()

    hace_30 = hoy - timedelta(days=30)

    clientes_totales = Compra.objects.values_list("cliente", flat=True).distinct().count()

    activos = (

        Compra.objects

        .filter(fecha_compra__gte=hace_30)

        .values_list("cliente", flat=True).distinct().count()

    )

    inactivos = max(clientes_totales - activos, 0)

    return JsonResponse({"activos": activos, "inactivos": inactivos})





@login_required

@require_http_methods(["GET"])

def api_admin_vendedores_estado(request):

    """

    Cuenta Activos vs Inactivos de vendedores para el donut.



    Modos:

      - Por defecto (legacy): usa `User.is_active` (habilitado/deshabilitado).

      - Con `presence=1`: usa presencia reciente mediante `last_login` en una

        ventana de tiempo (segundos) indicada por `window` (30..900; default 180).

        En este modo, "Activo" = visto en la ventana; "Inactivo" = resto.

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



    vendedor_id = request.GET.get("vendedor_id")

    presence = request.GET.get("presence") in {"1", "true", "True", "yes"}

    base = User.objects.filter(groups__name="Vendedores")



    if presence:

        try:

            window_seconds = int(request.GET.get("window", 180))

        except (TypeError, ValueError):

            window_seconds = 180

        window_seconds = max(30, min(900, window_seconds))

        limite = timezone.now() - timedelta(seconds=window_seconds)



        if vendedor_id:

            try:

                v = Vendedor.objects.select_related("usuario").get(pk=int(vendedor_id))

                u = v.usuario

                online = int(bool(u and u.is_active and u.last_login and u.last_login >= limite))

                offline = 1 - online

                return JsonResponse({"activos": online, "inactivos": offline, "scoped": True, "presence": True, "window": window_seconds})

            except (Vendedor.DoesNotExist, ValueError, TypeError):

                return JsonResponse({"activos": 0, "inactivos": 0, "scoped": True, "presence": True, "window": window_seconds})



                                                    

        online = base.filter(is_active=True, last_login__gte=limite).count()

        total = base.count()

        offline = max(0, total - online)

        return JsonResponse({"activos": online, "inactivos": offline, "scoped": False, "presence": True, "window": window_seconds})



                                                

    if vendedor_id:

        try:

            v = Vendedor.objects.select_related("usuario").get(pk=int(vendedor_id))

            activos = 1 if v.usuario and v.usuario.is_active else 0

            inactivos = 0 if activos == 1 else 1

            return JsonResponse({"activos": activos, "inactivos": inactivos, "scoped": True, "presence": False})

        except (Vendedor.DoesNotExist, ValueError, TypeError):

            return JsonResponse({"activos": 0, "inactivos": 0, "scoped": True, "presence": False})



    activos = base.filter(is_active=True).count()

    inactivos = base.filter(is_active=False).count()

    return JsonResponse({"activos": activos, "inactivos": inactivos, "scoped": False, "presence": False})





@login_required

@require_http_methods(["GET"])

def api_admin_usuarios_online(request):

    """Devuelve los IDs de usuarios activos en una ventana de tiempo reciente.



    Query params:

      - window: segundos de ventana (30..900, por defecto 180)

    Respuesta: { active_ids: [int, ...], window_seconds: int, generated_at: iso }

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



    try:

        window_seconds = int(request.GET.get("window", 180))

    except (TypeError, ValueError):

        window_seconds = 180

    window_seconds = max(30, min(900, window_seconds))



    limite = timezone.now() - timedelta(seconds=window_seconds)

    activos = list(

        User.objects.filter(is_active=True, last_login__gte=limite)

        .values_list("id", flat=True)

    )

    return JsonResponse({

        "active_ids": [int(pk) for pk in activos],

        "window_seconds": window_seconds,

        "generated_at": timezone.now().isoformat(),

    })





@login_required

@require_http_methods(["GET"])

def api_admin_top_productos_linea(request):

    """

    Serie por día de los TOP N productos (por total vendido) en los

    últimos `days` días. Opcionalmente filtrado por `vendedor_id`.

    Params:

      - vendedor_id (opcional)

      - days (opcional, default 30)

      - top_n (opcional, default 3)

    Respuesta:

      { labels: ["YYYY-MM-DD", ...], datasets: [{ label, data: [..] , product_id } ...] }

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



    vendedor_id = request.GET.get("vendedor_id")

    try:

        days = int(request.GET.get("days", 30))

        if days <= 0:

            days = 30

    except (TypeError, ValueError):

        days = 30

    try:

        top_n = int(request.GET.get("top_n", 3))

        if top_n <= 0:

            top_n = 3

    except (TypeError, ValueError):

        top_n = 3



    hoy = timezone.localdate()

    desde = hoy - timedelta(days=days - 1)



    ventas = Venta.objects.filter(fecha_venta__gte=desde, fecha_venta__lte=hoy)

    if vendedor_id:

        try:

            ventas = ventas.filter(vendedor_id=int(vendedor_id))

        except (TypeError, ValueError):

            ventas = ventas.none()



                                           

    top = (

        ventas

        .values("producto_id", "producto__nombre")

        .annotate(total=Sum("total"))

        .order_by("-total")[:top_n]

    )

    top_ids = [r["producto_id"] for r in top]

    top_names = {r["producto_id"]: (r["producto__nombre"] or f"Producto {r['producto_id']}") for r in top}



                             

    dias = [desde + timedelta(days=i) for i in range(days)]

    labels = [d.isoformat() for d in dias]



                                       

    datasets = []

    for pid in top_ids:

        serie = ventas.filter(producto_id=pid).values("fecha_venta").annotate(t=Sum("total"))

        mapa = {row["fecha_venta"].isoformat(): float(row["t"] or 0) for row in serie}

        data = [mapa.get(lbl, 0.0) for lbl in labels]

        datasets.append({

            "label": top_names.get(pid, f"Producto {pid}"),

            "data": data,

            "product_id": pid,

        })



    return JsonResponse({"labels": labels, "datasets": datasets})





                                                                



@login_required

@require_http_methods(["GET", "POST", "PUT", "PATCH", "DELETE"])

def api_admin_vendedores(request):

    """

    CRUD básico de usuarios-vendedores.

    GET:   lista (con filtros q y estado)

    POST:  crear (username, email?, password) -> añade al grupo 'Vendedores' y crea Vendedor

    PUT:   editar (id + campos)

    DEL:   toggle is_active (activar/desactivar)

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



                                  

    if request.method == "GET":

        estado = request.GET.get("estado")                                      

        q = (request.GET.get("q") or "").strip()



        usuarios = User.objects.all().order_by("id")

        if estado == "activo":

            usuarios = usuarios.filter(is_active=True)

        elif estado == "inactivo":

            usuarios = usuarios.filter(is_active=False)



        if q:

            usuarios = usuarios.filter(username__icontains=q) | usuarios.filter(email__icontains=q)



        vend_map = {u_id: v_id for v_id, u_id in Vendedor.objects.values_list("id", "usuario_id")}

        items = []

        for u in usuarios:

            es_vend = u.id in vend_map

            es_admin = bool(u.is_staff or u.is_superuser)

            rol = "Administrador" if es_admin else ("Vendedor" if es_vend else "Usuario")

            items.append({

                "id": u.id,

                "username": u.username,

                "email": u.email or "",

                "date_joined": u.date_joined.strftime("%Y-%m-%d"),

                "last_login": (u.last_login.strftime("%Y-%m-%d %H:%M") if u.last_login else "-"),

                "is_active": u.is_active,

                "es_vendedor": es_vend,

                "es_admin": es_admin,

                "is_self": u.id == request.user.id,

                "role": rol,

                "vendedor_id": vend_map.get(u.id),

                "password_hash": u.password if (request.user.is_superuser or request.user.is_staff) else "****",

            })

        return JsonResponse({"items": items})



    data = _json_body(request)



                                 

    if request.method == "POST":

        try:

            username = (data.get("username") or "").strip()

            email = (data.get("email") or "").strip()

            password = (data.get("password") or "").strip()

            if not username or not password:

                return _bad_request("username y password son obligatorios")

            if len(username) < 3:

                return _bad_request("username debe tener al menos 3 caracteres")

            if User.objects.filter(username=username).exists():

                return _bad_request("username ya existe")

                                                                

            if email:

                try:

                    validate_email(email)

                except ValidationError:

                    return _bad_request("email invalido")

                if User.objects.filter(email__iexact=email).exists():

                    return _bad_request("email ya esta en uso")

                                                                      

            try:

                validate_password(password)

            except ValidationError as e:

                return _bad_request(" ".join([str(m) for m in e.messages]))



            u = User.objects.create_user(username=username, email=email, password=password)

                                                            

            es_vendedor = bool(data.get("es_vendedor"))

            es_admin = bool(data.get("es_admin"))

            u.is_active = True

            if es_admin:

                u.is_staff = True

            u.save()

            if es_vendedor:

                grupo, _ = Group.objects.get_or_create(name="Vendedores")

                u.groups.add(grupo)

                                                                                 
                                                                              
                                                                             
                                                                    
                try:
                    Vendedor.objects.get_or_create(usuario=u)
                except Exception:
                    try:
                        with connection.cursor() as cur:
                                                                    
                                                                             
                            cur.execute(
                                "INSERT INTO core_vendedor (usuario_id, telefono, direccion, fecha_ingreso, umbral_critico) "
                                "VALUES (%s, %s, %s, CURRENT_DATE, %s)",
                                [u.id, "", "", 5],
                            )
                    except Exception as e2:
                        return _bad_request(f"server: {e2}")

                                                                                

                try:

                    if email:

                        asunto = "Bienvenido como vendedor a EpicAnimes!"

                        panel_url = request.build_absolute_uri(reverse("dashboard_vendedor"))

                        cuerpo = (

                            f"Hola {username},\n\n"

                            "Bienvenido a EpicAnimes! Tu cuenta de vendedor ha sido creada y aprobada.\n"

                            "Desde ahora puedes ingresar a tu panel para publicar y administrar productos, revisar ventas y gestionar tu catálogo.\n\n"

                            f"Panel de vendedor: {panel_url}\n\n"

                            "Si tienes dudas o necesitas ayuda, responde a este correo y nuestro equipo te apoyará.\n\n"

                            "¡éxito en tus ventas!"

                        )

                        send_mail(asunto, cuerpo, settings.DEFAULT_FROM_EMAIL, [email], fail_silently=False)

                except Exception:

                                                             

                    pass

            return JsonResponse({"ok": True, "id": u.id})

        except Exception as e:

            return _bad_request(f"server: {e}")



                                  

    if request.method in ("PUT", "PATCH"):

        uid = data.get("id")

        if not uid:

            return _bad_request("id requerido")

        try:

            u = User.objects.get(id=uid)

        except User.DoesNotExist:

            return _bad_request("usuario no existe")



                                                                                       

        if data.get("reset_password"):

            tmp = User.objects.make_random_password(length=10)

            u.set_password(tmp)

            u.save()

            return JsonResponse({"ok": True, "temp_password": tmp})



        username = data.get("username")

        email = data.get("email")

        password = data.get("password")

        is_active = data.get("is_active")

        es_vendedor = data.get("es_vendedor", True)



        if username:

            if User.objects.exclude(id=u.id).filter(username=username).exists():

                return HttpResponseBadRequest("username ya en uso")

            u.username = username

        if email is not None:

            email_s = (email or "").strip()

            if email_s:

                try:

                    validate_email(email_s)

                except ValidationError:

                    return _bad_request("email invalido")

                if User.objects.exclude(id=u.id).filter(email__iexact=email_s).exists():

                    return _bad_request("email ya esta en uso")

            u.email = email_s

        if isinstance(is_active, bool):

            u.is_active = is_active

        if password:

            try:

                validate_password(password, user=u)

            except ValidationError as e:

                return _bad_request(" ".join([str(m) for m in e.messages]))

            u.set_password(password)

        u.save()



        grupo, _ = Group.objects.get_or_create(name="Vendedores")

        if es_vendedor:

                                                                             
            try:
                Vendedor.objects.get_or_create(usuario=u)
            except Exception:
                try:
                    with connection.cursor() as cur:
                        cur.execute(
                            "INSERT INTO core_vendedor (usuario_id, telefono, direccion, fecha_ingreso, umbral_critico) "
                            "VALUES (%s, %s, %s, CURRENT_DATE, %s)",
                            [u.id, "", "", 5],
                        )
                except Exception as e2:
                    return _bad_request(f"server: {e2}")

            u.groups.add(grupo)

        else:

            Vendedor.objects.filter(usuario=u).delete()

            u.groups.remove(grupo)



        return JsonResponse({"ok": True})



                                                

    if request.method == "DELETE":

        uid = data.get("id")

        if not uid:

            return _bad_request("id requerido")

        try:

            u = User.objects.get(id=uid)

        except User.DoesNotExist:

            return _bad_request("usuario no existe")



        if data.get("eliminar"):

            if u.id == request.user.id:

                return _bad_request("No puedes eliminar tu propio usuario.")

            if u.is_superuser and not request.user.is_superuser:

                return HttpResponseForbidden("No puedes eliminar este usuario.")

            Vendedor.objects.filter(usuario=u).delete()

            u.delete()

            return JsonResponse({"ok": True, "eliminado": True})



        u.is_active = not u.is_active

        u.save()

        return JsonResponse({"ok": True, "is_active": u.is_active})



    return HttpResponseNotAllowed(["GET", "POST", "PUT", "PATCH", "DELETE"])





                                                              

                                                   

                                                              



@login_required

def redireccion_usuario(request):

    """Redirige al usuario al panel correspondiente según su rol."""
    user = request.user

    if user.is_superuser or user.is_staff:

        return redirect("dashboard_administrador")

    if user.groups.filter(name="Vendedores").exists():

        return redirect("dashboard_vendedor")

    return redirect("index")



                                                                         



@login_required

@require_http_methods(["GET", "PUT", "PATCH"])

def api_admin_postulaciones(request):

    """Entrega las postulaciones de vendedores en formato JSON."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")

    if request.method == "GET":

        qs = PostulacionVendedor.objects.all().order_by("-fecha_envio")

        q = (request.GET.get("q") or "").strip().lower()

        estado = (request.GET.get("estado") or "").strip().lower()

        desde_s = request.GET.get("from")

        hasta_s = request.GET.get("to")

        if estado in {"nuevo", "contactado", "archivado"}:

            qs = qs.filter(estado=estado)

        if q:

            qs = qs.filter(Q(nombre__icontains=q) | Q(email__icontains=q) | Q(mensaje__icontains=q))

        try:

            d = parse_date(desde_s) if desde_s else None

            if d:

                qs = qs.filter(fecha_envio__date__gte=d)

        except Exception:

            pass

        try:

            h = parse_date(hasta_s) if hasta_s else None

            if h:

                qs = qs.filter(fecha_envio__date__lte=h)

        except Exception:

            pass

        items = [

            {

                "id": p.id,

                "nombre": p.nombre,

                "email": p.email,

                "telefono": p.telefono or "",

                "tienda": p.tienda or "",

                "instagram": p.instagram or "",

                "mensaje": p.mensaje or "",

                "notas": p.notas or "",

                "fecha_envio": p.fecha_envio.strftime("%Y-%m-%d %H:%M"),

                "estado": p.estado,

            }

            for p in qs

        ]

        return JsonResponse({"items": items})



    data = _json_body(request)

    pid = data.get("id")

    nuevo_estado = (data.get("estado") or "").strip().lower()

    notas = data.get("notas")

    if not pid and notas is None:

        return HttpResponseBadRequest("id requerido")

    try:

        p = PostulacionVendedor.objects.get(pk=int(pid))

    except (PostulacionVendedor.DoesNotExist, ValueError, TypeError):

        return JsonResponse({"error": "not_found"}, status=404)

    updates = []

    if nuevo_estado in {"nuevo", "contactado", "archivado"}:

        p.estado = nuevo_estado

        updates.append("estado")

    if isinstance(notas, str):

        p.notas = notas

        updates.append("notas")

    if not updates:

        return HttpResponseBadRequest("sin cambios")

    p.save(update_fields=updates)

    return JsonResponse({"ok": True, "id": p.id, "estado": p.estado})





                                                                



@login_required

@require_http_methods(["GET"])

def export_admin_postulaciones_csv(request):

    """Genera un CSV con las postulaciones recibidas."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")

    import csv

    resp = HttpResponse(content_type='text/csv; charset=utf-8')

    resp["Content-Disposition"] = 'attachment; filename="postulaciones.csv"'

    writer = csv.writer(resp)

    writer.writerow(["id","nombre","email","telefono","tienda","instagram","mensaje","notas","fecha_envio","estado"])

    qs = PostulacionVendedor.objects.all().order_by("-fecha_envio")

    q = (request.GET.get("q") or "").strip().lower()

    estado = (request.GET.get("estado") or "").strip().lower()

    desde_s = request.GET.get("from")

    hasta_s = request.GET.get("to")

    if estado in {"nuevo","contactado","archivado"}: qs = qs.filter(estado=estado)

    if q: qs = qs.filter(Q(nombre__icontains=q)|Q(email__icontains=q)|Q(mensaje__icontains=q))

    try:

        d=parse_date(desde_s) if desde_s else None

        if d: qs=qs.filter(fecha_envio__date__gte=d)

    except Exception: pass

    try:

        h=parse_date(hasta_s) if hasta_s else None

        if h: qs=qs.filter(fecha_envio__date__lte=h)

    except Exception: pass

    for p in qs:

        writer.writerow([p.id, p.nombre, p.email, p.telefono or '', p.tienda or '', p.instagram or '', (p.mensaje or '').replace('\n',' ').strip(), (p.notas or '').replace('\n',' ').strip(), p.fecha_envio.isoformat(sep=' '), p.estado])

    return resp





@login_required

@require_http_methods(["GET"])

def export_admin_ventas_csv(request):

    """Exporta las ventas en formato CSV para administradores."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")

    import csv

    desde_s = request.GET.get('from') or request.GET.get('start')

    hasta_s = request.GET.get('to') or request.GET.get('end')

    days = request.GET.get('days')

    qs = Venta.objects.select_related('vendedor__usuario','producto')

    if desde_s:

        try:

            d = parse_date(desde_s)

            if d: qs = qs.filter(fecha_venta__gte=d)

        except Exception: pass

    if hasta_s:

        try:

            h = parse_date(hasta_s)

            if h: qs = qs.filter(fecha_venta__lte=h)

        except Exception: pass

    if not desde_s and not hasta_s and days:

        try:

            n = int(days)

            n = max(1, min(365, n))

            hoy = timezone.localdate()

            qs = qs.filter(fecha_venta__gte=hoy - timedelta(days=n))

        except Exception: pass

    resp = HttpResponse(content_type='text/csv; charset=utf-8')

    resp["Content-Disposition"] = 'attachment; filename="ventas.csv"'

    writer = csv.writer(resp)

    writer.writerow(["fecha","vendedor","producto","cantidad","total"])

    for v in qs.order_by('fecha_venta','vendedor_id'):

        writer.writerow([v.fecha_venta.isoformat(), getattr(getattr(v.vendedor,'usuario',None),'username',''), v.producto.nombre, v.cantidad, f"{v.total}"])

    return resp





@login_required

@require_http_methods(["GET"])

def export_vendedor_inventario_csv(request):

                   

    """Construye un CSV con el inventario del vendedor."""
    if not request.user.groups.filter(name="Vendedores").exists():

        return HttpResponseForbidden("Solo vendedores")

    try:

        vend = Vendedor.objects.get(usuario=request.user)

    except Vendedor.DoesNotExist:

        return HttpResponseForbidden("Perfil vendedor requerido")

    import csv

    resp = HttpResponse(content_type='text/csv; charset=utf-8')

    resp["Content-Disposition"] = 'attachment; filename="inventario.csv"'

    writer = csv.writer(resp)

    writer.writerow(["nombre","marca","calidad","categoria","precio","existencias","fecha_ingreso","descripcion"])

    for p in Producto.objects.filter(vendedor=vend).order_by('nombre'):

        writer.writerow([p.nombre, p.marca, p.calidad, p.categoria, f"{p.precio}", int(p.existencias or 0), (p.fecha_ingreso.isoformat() if p.fecha_ingreso else ''), (p.descripcion or '').replace('\n',' ')])

    return resp





@login_required

@require_http_methods(["GET"])

def export_vendedor_ventas_csv(request):

    """Exporta las ventas del vendedor en CSV."""
    if not request.user.groups.filter(name="Vendedores").exists():

        return HttpResponseForbidden("Solo vendedores")

    try:

        vend = Vendedor.objects.get(usuario=request.user)

    except Vendedor.DoesNotExist:

        return HttpResponseForbidden("Perfil vendedor requerido")

    desde_s = request.GET.get('from')

    hasta_s = request.GET.get('to')

    days = request.GET.get('days')

    qs = Venta.objects.filter(vendedor=vend).select_related('producto')

    if desde_s:

        d = parse_date(desde_s)

        if d: qs = qs.filter(fecha_venta__gte=d)

    if hasta_s:

        h = parse_date(hasta_s)

        if h: qs = qs.filter(fecha_venta__lte=h)

    if not desde_s and not hasta_s and days:

        try:

            n = int(days)

            n = max(1, min(365, n))

            hoy = timezone.localdate()

            qs = qs.filter(fecha_venta__gte=hoy - timedelta(days=n))

        except Exception:

            pass

    import csv

    resp = HttpResponse(content_type='text/csv; charset=utf-8')

    resp["Content-Disposition"] = 'attachment; filename="ventas_vendedor.csv"'

    w = csv.writer(resp)

    w.writerow(["fecha","producto","cantidad","total"])

    for v in qs.order_by('fecha_venta'):

        w.writerow([v.fecha_venta.isoformat(), v.producto.nombre, v.cantidad, f"{v.total}"])

    return resp





@login_required

@require_http_methods(["POST"])

def api_vendedor_importar(request):

                   

    """Permite cargar productos en lote para el vendedor autenticado."""
    if not request.user.groups.filter(name="Vendedores").exists():

        return HttpResponseForbidden("Solo vendedores")

    try:

        vend = Vendedor.objects.get(usuario=request.user)

    except Vendedor.DoesNotExist:

        return HttpResponseForbidden("Perfil vendedor requerido")

    f = request.FILES.get('file') or request.FILES.get('csv')

    if not f:

        return HttpResponseBadRequest("Archivo 'file' CSV requerido")

    import csv, io

    try:

        data = f.read().decode('utf-8', errors='ignore')

        sample = data[:1024]

        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")

    except Exception:

        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(data), dialect=dialect)

    created = 0

    for row in reader:

        nombre = (row.get('nombre') or row.get('Nombre') or '').strip()

        if not nombre:

            continue

        marca = (row.get('marca') or row.get('Marca') or '').strip()

        calidad = (row.get('calidad') or row.get('Calidad') or '').strip()

        categoria = (row.get('categoria') or row.get('Categoría') or row.get('Categoria') or '').strip()

        try:

            precio = Decimal(str(row.get('precio') or row.get('Precio') or '0')).quantize(Decimal('0.01'))

        except Exception:

            precio = Decimal('0.00')

        try:

            exist = int((row.get('existencias') or row.get('stock') or row.get('Existencias') or '0').strip() or 0)

        except Exception:

            exist = 0

        fecha = parse_date((row.get('fecha_ingreso') or row.get('Fecha') or '').strip() or '') or timezone.localdate()

        desc = (row.get('descripcion') or row.get('Descripción') or '').strip()

        Producto.objects.create(

            vendedor=vend,

            nombre=nombre,

            marca=marca,

            calidad=calidad,

            categoria=categoria,

            precio=precio,

            existencias=exist,

            fecha_ingreso=fecha,

            descripcion=desc,

        )

        created += 1

    return JsonResponse({"ok": True, "creados": created})



                       

                       

                       

@login_required

@require_http_methods(["GET"])

def export_admin_postulaciones_xlsx(request):

    """Genera un archivo XLSX de postulaciones."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo administradores")

    try:

        from openpyxl import Workbook

    except Exception:

        return HttpResponseBadRequest("Falta dependencia 'openpyxl'")

    from io import BytesIO

    wb = Workbook()

    ws = wb.active

    ws.title = "Postulaciones"

    ws.append(["ID","Nombre","Email","Teléfono","Tienda","Instagram/Web","Mensaje","Fecha","Estado","Notas"])

    q = (request.GET.get("q") or "").strip().lower()

    estado = (request.GET.get("estado") or "").strip().lower()

    desde_s = request.GET.get("from")

    hasta_s = request.GET.get("to")

    qs = PostulacionVendedor.objects.all()

    if estado in {"nuevo","contactado","archivado"}:

        qs = qs.filter(estado=estado)

    if q:

        qs = qs.filter(Q(nombre__icontains=q) | Q(email__icontains=q) | Q(mensaje__icontains=q))

    try:

        d = parse_date(desde_s) if desde_s else None

        if d:

            qs = qs.filter(fecha_envio__date__gte=d)

    except Exception:

        pass

    try:

        h = parse_date(hasta_s) if hasta_s else None

        if h:

            qs = qs.filter(fecha_envio__date__lte=h)

    except Exception:

        pass

    for p in qs.order_by("-fecha_envio"):

        ws.append([

            p.id,

            p.nombre,

            p.email,

            p.telefono or "",

            p.tienda or "",

            p.instagram or "",

            (p.mensaje or "")[:500],

            p.fecha_envio.isoformat(sep=" "),

            p.estado,

            (p.notas or "")[:500],

        ])

    bio = BytesIO(); wb.save(bio); bio.seek(0)

    resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    resp["Content-Disposition"] = 'attachment; filename="postulaciones.xlsx"'

    return resp





@login_required

@require_http_methods(["GET"])

def export_admin_ventas_xlsx(request):

    """Exporta las ventas en formato XLSX."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo administradores")

    try:

        from openpyxl import Workbook

    except Exception:

        return HttpResponseBadRequest("Falta dependencia 'openpyxl'")

    desde_s = request.GET.get('from') or request.GET.get('start')

    hasta_s = request.GET.get('to') or request.GET.get('end')

    days = request.GET.get('days')

    qs = Venta.objects.select_related('vendedor__usuario','producto')

    if desde_s:

        try:

            d = parse_date(desde_s)

            if d: qs = qs.filter(fecha_venta__gte=d)

        except Exception:

            pass

    if hasta_s:

        try:

            h = parse_date(hasta_s)

            if h: qs = qs.filter(fecha_venta__lte=h)

        except Exception:

            pass

    if not desde_s and not hasta_s and days:

        try:

            n = int(days)

            n = max(1, min(365, n))

            hoy = timezone.localdate()

            qs = qs.filter(fecha_venta__gte=hoy - timedelta(days=n))

        except Exception:

            pass

    from io import BytesIO

    wb = Workbook(); ws = wb.active; ws.title = "Ventas"

    ws.append(["Fecha","Vendedor","Producto","Cantidad","Total"])

    for v in qs.order_by('fecha_venta','vendedor_id'):

        ws.append([

            v.fecha_venta.isoformat(),

            getattr(getattr(v.vendedor,'usuario',None),'username',''),

            v.producto.nombre,

            v.cantidad,

            float(v.total),

        ])

    bio = BytesIO(); wb.save(bio); bio.seek(0)

    resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    resp["Content-Disposition"] = 'attachment; filename="ventas.xlsx"'

    return resp





@login_required

@require_http_methods(["GET"])

def export_vendedor_inventario_xlsx(request):

    """Genera un XLSX del inventario del vendedor."""
    if not request.user.groups.filter(name="Vendedores").exists():

        return HttpResponseForbidden("Solo vendedores")

    try:

        vend = Vendedor.objects.get(usuario=request.user)

    except Vendedor.DoesNotExist:

        return HttpResponseForbidden("Perfil vendedor requerido")

    try:

        from openpyxl import Workbook

    except Exception:

        return HttpResponseBadRequest("Falta dependencia 'openpyxl'")

    from io import BytesIO

    wb = Workbook(); ws = wb.active; ws.title = "Inventario"

    ws.append(["Nombre","Marca","Calidad","Categoría","Precio","Existencias","Fecha ingreso","Descripción"])

    for p in Producto.objects.filter(vendedor=vend).order_by('nombre'):

        ws.append([

            p.nombre,

            p.marca,

            p.calidad,

            p.categoria,

            float(p.precio),

            int(p.existencias or 0),

            (p.fecha_ingreso.isoformat() if p.fecha_ingreso else ''),

            (p.descripcion or ''),

        ])

    bio = BytesIO(); wb.save(bio); bio.seek(0)

    resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    resp["Content-Disposition"] = 'attachment; filename="inventario.xlsx"'

    return resp





@login_required

@require_http_methods(["GET"])

def export_vendedor_ventas_xlsx(request):

    """Exporta las ventas del vendedor en XLSX."""
    if not request.user.groups.filter(name="Vendedores").exists():

        return HttpResponseForbidden("Solo vendedores")

    try:

        vend = Vendedor.objects.get(usuario=request.user)

    except Vendedor.DoesNotExist:

        return HttpResponseForbidden("Perfil vendedor requerido")

    try:

        from openpyxl import Workbook

    except Exception:

        return HttpResponseBadRequest("Falta dependencia 'openpyxl'")

    desde_s = request.GET.get('from')

    hasta_s = request.GET.get('to')

    days = request.GET.get('days')

    qs = Venta.objects.filter(vendedor=vend).select_related('producto')

    if desde_s:

        d = parse_date(desde_s)

        if d: qs = qs.filter(fecha_venta__gte=d)

    if hasta_s:

        h = parse_date(hasta_s)

        if h: qs = qs.filter(fecha_venta__lte=h)

    if not desde_s and not hasta_s and days:

        try:

            n = int(days)

            n = max(1, min(365, n))

            hoy = timezone.localdate()

            qs = qs.filter(fecha_venta__gte=hoy - timedelta(days=n))

        except Exception:

            pass

    from io import BytesIO

    wb = Workbook(); ws = wb.active; ws.title = "Ventas"

    ws.append(["Fecha","Producto","Cantidad","Total"])

    for v in qs.order_by('fecha_venta'):

        ws.append([v.fecha_venta.isoformat(), v.producto.nombre, v.cantidad, float(v.total)])

    bio = BytesIO(); wb.save(bio); bio.seek(0)

    resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    resp["Content-Disposition"] = 'attachment; filename="ventas_vendedor.xlsx"'

    return resp





                            

                        

                            

@login_required

@require_http_methods(["POST"])

def api_vendedor_importar_excel(request):

    """Procesa archivos Excel para crear o actualizar productos."""
    if not request.user.groups.filter(name="Vendedores").exists():

        return HttpResponseForbidden("Solo vendedores")

    try:

        vend = Vendedor.objects.get(usuario=request.user)

    except Vendedor.DoesNotExist:

        return HttpResponseForbidden("Perfil vendedor requerido")

    f = request.FILES.get('file')

    if not f:

        return HttpResponseBadRequest("Archivo Excel requerido")

    try:

        from openpyxl import load_workbook

        from io import BytesIO

        content = f.read()

        wb = load_workbook(filename=BytesIO(content), data_only=True)

        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        if not rows:

            return JsonResponse({"ok": True, "creados": 0})

        def _norm(s):

            if not s:

                return ''

            import unicodedata

            s = str(s)

            s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))

            return s.strip()

        headers = [(_norm(h) or '').lower().replace(' ', '_') for h in (rows[0] or [])]

        created = 0

        for r in rows[1:]:

            row = {}

            for idx, val in enumerate(r or []):

                key = headers[idx] if idx < len(headers) else f'col_{idx}'

                row[key] = val if val is not None else ''

            nombre = _norm(row.get('nombre') or row.get('name'))

            if not nombre:

                continue

            marca = _norm(row.get('marca'))

            calidad = _norm(row.get('calidad'))

            categoria = _norm(row.get('categoria') or row.get('categoria_'))

            try:

                precio = Decimal(str(row.get('precio') or '0')).quantize(Decimal('0.01'))

            except Exception:

                precio = Decimal('0.00')

            try:

                exist = int(str(row.get('existencias') or row.get('stock') or '0'))

            except Exception:

                exist = 0

            fecha_raw = str(row.get('fecha_ingreso') or '')

            fecha = parse_date(fecha_raw) or timezone.localdate()

            desc = str(row.get('descripcion') or '')

            Producto.objects.create(

                vendedor=vend,

                nombre=nombre,

                marca=marca,

                calidad=calidad,

                categoria=categoria,

                precio=precio,

                existencias=exist,

                fecha_ingreso=fecha,

                descripcion=desc,

            )

            created += 1

        return JsonResponse({"ok": True, "creados": created})

    except Exception:

        return HttpResponseBadRequest("No se pudo procesar el Excel")













class CoreLoginView(LoginView):

    """Renderiza el formulario de inicio de sesión con verificación OTP."""
    template_name = 'registration/login.html'

    authentication_form = TwoFactorLoginForm


class CustomPasswordResetView(PasswordResetView):
    """Personaliza el restablecimiento de contraseña mostrando mensajes claros."""
    template_name = "registration/password_reset_form.html"
    email_template_name = "registration/password_reset_email.html"
    subject_template_name = "registration/password_reset_subject.txt"
    success_url = reverse_lazy("password_reset")

    def form_valid(self, form):
        """Valida que exista un usuario asociado y entrega mensajes consistentes."""
        email = form.cleaned_data.get("email", "").strip()
        users = list(form.get_users(email))
        if not users:
            messages.error(self.request, "No encontramos una cuenta asociada a ese correo.")
            form.add_error("email", "No encontramos este correo en EpicAnimes.")
            return self.render_to_response(self.get_context_data(form=form))
        response = super().form_valid(form)
        messages.success(self.request, "Si el correo existe, te enviamos un enlace para restablecer tu contraseña.")
        return response

    def form_invalid(self, form):
        """Refuerza los mensajes de error cuando el correo ingresado es inválido."""
        if form.errors.get("email"):
            messages.error(self.request, "Revisa el correo ingresado.")
        return super().form_invalid(form)




@login_required

def VistaCarrito(request):

    """Muestra el carrito del usuario junto con la información de checkout."""
    cart = _get_cart(request)

    items, total, carrito_sin_fallos = _build_cart_items(cart)

    rol_actual = obtener_rol_usuario(request.user)

    es_comprador = rol_actual == "comprador"

    paypal_configurado, paypal_error = paypal_is_configured()



    perfil_cliente = None

    if request.user.is_authenticated:

        perfil_cliente = getattr(request.user, "perfil_cliente", None)

        if perfil_cliente is None:

            perfil_cliente = PerfilCliente.objects.filter(user=request.user).first()



    checkout_prefill = request.session.get("checkout_info_prefill", {}).copy()

    if request.user.is_authenticated:

        if perfil_cliente:

            checkout_prefill.setdefault("nombre", perfil_cliente.nombre or request.user.get_full_name() or request.user.username)

            checkout_prefill.setdefault("email", perfil_cliente.email or request.user.email or "")

            checkout_prefill.setdefault("telefono", perfil_cliente.telefono or "")

            checkout_prefill.setdefault("direccion", perfil_cliente.direccion or "")

            checkout_prefill.setdefault("ciudad", perfil_cliente.ciudad or "")

        else:

            checkout_prefill.setdefault("nombre", request.user.get_full_name() or request.user.username)

            checkout_prefill.setdefault("email", request.user.email or "")

    checkout_prefill.setdefault("telefono", checkout_prefill.get("telefono", ""))

    checkout_prefill.setdefault("direccion", checkout_prefill.get("direccion", ""))

    checkout_prefill.setdefault("ciudad", checkout_prefill.get("ciudad", ""))

    checkout_prefill.setdefault("notas", checkout_prefill.get("notas", ""))



    puede_pagar = es_comprador and carrito_sin_fallos and bool(items)

    paypal_summary = paypal_conversion_summary(total)

    contexto = {

        "items": items,

        "total": total,

        "carrito_sin_fallos": carrito_sin_fallos,

        "paypal_client_id": getattr(settings, "PAYPAL_CLIENT_ID", ""),

        **paypal_summary,

        "paypal_enabled": paypal_configurado,

        "paypal_error": paypal_error,

        "cart_count": _cart_count(cart),

        "rol_usuario": rol_actual,

        "puede_comprar": es_comprador,

        "puede_pagar": puede_pagar,

        "puede_pagar_paypal": puede_pagar and paypal_configurado,

        "checkout_prefill": checkout_prefill,

    }

    return render(request, "public/carrito.html", contexto)





@login_required

@require_http_methods(["POST"])

def agregar_al_carrito(request, producto_id):

    """Agrega un producto al carrito del usuario."""
    if obtener_rol_usuario(request.user) != "comprador":

        messages.error(request, "Tu rol no permite comprar en la tienda.")

        return redirect("index")

    producto = get_object_or_404(Producto, pk=producto_id)

    try:

        cantidad = int(request.POST.get("cantidad", 1))

    except (TypeError, ValueError):

        cantidad = 1

    cantidad = max(1, cantidad)

    cart = _get_cart(request)

    actual = cart.get(str(producto.id), 0)

    if actual + cantidad > producto.existencias:

        messages.error(request, "No hay stock suficiente del producto seleccionado.", extra_tags="critico")

        return redirect("index")

    cart[str(producto.id)] = actual + cantidad

    _save_cart(request, cart)

    messages.success(request, f"{producto.nombre} agregado al carrito.")

    return redirect("carrito")





@login_required

@require_http_methods(["POST"])

def actualizar_carrito(request, producto_id):

    """Actualiza las cantidades de un producto dentro del carrito."""
    is_json = request.headers.get("x-requested-with") == "XMLHttpRequest" or (request.content_type or "").startswith("application/json")

    if obtener_rol_usuario(request.user) != "comprador":

        if is_json:

            return JsonResponse({"ok": False, "error": "Tu rol no permite modificar el carrito."}, status=400)

        messages.error(request, "Tu rol no permite modificar el carrito.")

        return redirect("carrito")

    cart = _get_cart(request)

    key = str(producto_id)

    if key not in cart:

        if is_json:

            return JsonResponse({"ok": False, "error": "El producto no esta en tu carrito."}, status=404)

        messages.error(request, "El producto no esta en tu carrito.")

        return redirect("carrito")



    if is_json:

        try:

            payload = json.loads(request.body or "{}")

        except json.JSONDecodeError:

            payload = {}

        cantidad_raw = payload.get("cantidad", cart[key])

    else:

        cantidad_raw = request.POST.get("cantidad", cart[key])

    try:

        cantidad = int(cantidad_raw)

    except (TypeError, ValueError):

        cantidad = cart[key]



    if cantidad <= 0:

        cart.pop(key, None)

    else:

        producto = get_object_or_404(Producto, pk=producto_id)

        if cantidad > producto.existencias:

            if is_json:

                return JsonResponse({"ok": False, "error": "Stock insuficiente para el producto."}, status=400)

            messages.error(request, "Stock insuficiente para el producto.")

            return redirect("carrito")

        cart[key] = cantidad

    _save_cart(request, cart)



    if is_json:

        items, total, puede_pagar = _build_cart_items(cart)

        subtotal = Decimal("0")

        sin_stock = False

        stock_disponible = None

        for item in items:

            if item["producto"].id == int(producto_id):

                subtotal = item["subtotal"]

                sin_stock = item["sin_stock"]

                stock_disponible = item["producto"].existencias

                break

        return JsonResponse({

            "ok": True,

            "cantidad": cart.get(key, 0),

            "subtotal": f"{subtotal:.0f}",

            "subtotal_raw": f"{subtotal:.2f}",

            "total": f"{total:.0f}",

            "total_raw": f"{total:.2f}",

            "sin_stock": sin_stock,

            "stock_disponible": stock_disponible,

            "puede_pagar": puede_pagar,

            "cart_count": _cart_count(cart),

        })



    messages.success(request, "Carrito actualizado.")

    return redirect("carrito")

@login_required

@require_http_methods(["POST"])

def eliminar_del_carrito(request, producto_id):

    """Elimina un producto del carrito."""
    is_json = request.headers.get("x-requested-with") == "XMLHttpRequest" or (request.content_type or "").startswith("application/json")

    cart = _get_cart(request)

    key = str(producto_id)

    if key in cart:

        cart.pop(key)

        _save_cart(request, cart)

        if is_json:

            items, total, puede_pagar = _build_cart_items(cart)

            return JsonResponse({

                "ok": True,

                "total_raw": f"{total:.2f}",

                "cart_count": _cart_count(cart),

                "puede_pagar": puede_pagar,

                "items_restantes": len(items),

            })

        messages.success(request, "Producto eliminado del carrito.")

    else:

        if is_json:

            return JsonResponse({"ok": False, "error": "El producto ya no estaba en tu carrito."}, status=404)

        messages.info(request, "El producto ya no estaba en tu carrito.")

    return redirect("carrito")





@login_required

@require_http_methods(["POST"])

def paypal_crear_orden(request):

    """Crea una orden de PayPal para el contenido del carrito."""
    if obtener_rol_usuario(request.user) != "comprador":

        return JsonResponse({"ok": False, "error": "Tu rol no permite comprar en la tienda."}, status=403)



    content_type = request.content_type or ""

    cart = _get_cart(request)

    print("[paypal_crear_orden] usuario=", getattr(request.user, "id", None), "cart_keys=", list(cart.keys()))

    if content_type.startswith("application/json"):

        try:

            payload = json.loads(request.body or "{}")

        except json.JSONDecodeError:

            payload = {}

        datos_cliente = payload.get("datos_cliente") or {}

    else:

        datos_cliente = {

            "nombre": request.POST.get("nombre") or "",

            "email": request.POST.get("email") or "",

            "telefono": request.POST.get("telefono") or "",

            "direccion": request.POST.get("direccion") or "",

            "ciudad": request.POST.get("ciudad") or "",

            "notas": request.POST.get("notas") or "",

        }



    try:

        datos_normalizados = _resolver_datos_cliente(request, datos_cliente)

        _lineas, total = _calcular_lineas_y_total(cart, lock=False)

    except CarritoError as exc:

        print("[paypal_crear_orden] error datos/carrito:", exc)

        return JsonResponse({"ok": False, "error": str(exc)}, status=400)



    moneda_paypal = getattr(settings, "PAYPAL_CURRENCY", "CLP")

    try:
        total, order_total, conversion_rate, moneda_paypal, order_currency = _calcular_totales_paypal(total)
    except PayPalError as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)

    print("[paypal_crear_orden] total_normalizado=", total, "moneda=", moneda_paypal)



    shipping = {

        "name": {"full_name": datos_normalizados["nombre"][:300]},

        "address": {

            "address_line_1": datos_normalizados["direccion"][:300] or "Direccion pendiente",

            "admin_area_1": "RM",

            "admin_area_2": datos_normalizados["ciudad"][:120] or "Santiago",

            "postal_code": "8320000",

            "country_code": "CL",

        },

    }

    try:

        reference = f"ORD-{request.user.id}-{timezone.now().strftime('%Y%m%d%H%M%S')}"

        print(

            "[paypal_crear_orden] order_total=",

            order_total,

            "order_currency=",

            order_currency,

            "conversion_rate=",

            conversion_rate,

        )

        order_id = paypal_create_order(order_total, order_currency, shipping=shipping, reference=reference)

    except PayPalError as exc:

        print("[paypal_crear_orden] error paypal:", exc)

        return JsonResponse({"ok": False, "error": str(exc)}, status=400)



    request.session["checkout_info_prefill"] = datos_normalizados.copy()

    request.session.modified = True



    return JsonResponse({"ok": True, "orderID": order_id})





@login_required

@require_http_methods(["POST"])

def finalizar_compra(request):

    """Valida el carrito y crea la orden de compra definitiva."""
    content_type = request.content_type or ""

    is_json = content_type.startswith("application/json") or request.headers.get("x-requested-with") == "XMLHttpRequest"

    referencia = None

    datos_cliente = {}

    if content_type.startswith("application/json"):

        try:

            payload = json.loads(request.body or "{}")

        except json.JSONDecodeError:

            payload = {}

        referencia = payload.get("paypal_order_id") or payload.get("orderID")

        datos_cliente = payload.get("datos_cliente") or {}

    else:

        referencia = request.POST.get("paypal_order_id")

        datos_cliente = {

            "nombre": request.POST.get("nombre") or "",

            "email": request.POST.get("email") or "",

            "telefono": request.POST.get("telefono") or "",

            "direccion": request.POST.get("direccion") or "",

            "ciudad": request.POST.get("ciudad") or "",

            "notas": request.POST.get("notas") or "",

        }

    if not referencia:

        mensaje_error = "No se recibio la confirmacion de PayPal. Intenta nuevamente."

        if is_json:

            return JsonResponse({"ok": False, "error": mensaje_error}, status=400)

        messages.error(request, mensaje_error)

        return redirect("carrito")



    try:

        total = _procesar_compra(request, referencia_pago=referencia, datos_cliente=datos_cliente)

    except CarritoError as exc:

        if is_json:

            return JsonResponse({"ok": False, "error": str(exc)}, status=400)

        messages.error(request, str(exc))

        return redirect("carrito")

    except Exception:

        if is_json:

            return JsonResponse({"ok": False, "error": "Ocurrio un error al procesar la compra."}, status=500)

        messages.error(request, "Ocurrio un error al procesar la compra.")

        return redirect("carrito")



    if is_json:

        return JsonResponse({"ok": True, "redirect": reverse("carrito_gracias"), "total": f"{total:.2f}"})



    messages.success(request, "Compra realizada correctamente.")

    return redirect("carrito_gracias")





@login_required

def carrito_gracias(request):

    """Muestra la página de confirmación después de pagar."""
    total = request.session.pop("ultimo_total", None)

    checkout_info = request.session.pop("ultimo_checkout_info", None)

    return render(request, "public/carrito_gracias.html", {"total": total, "checkout_info": checkout_info})








                                                              

                                       

                                                   

                                                              



@login_required

def VistaVendedor(request):

    """Renderiza el panel privado para los vendedores."""
    vendedor = Vendedor.objects.filter(usuario=request.user).first()

    productos = Producto.objects.none()

    if vendedor:

        productos = Producto.objects.filter(vendedor=vendedor)

    form_data = {}



    if request.method == "POST":

        form_data = request.POST.dict()



        nombre = form_data.get("nombre", "").strip()

        descripcion = form_data.get("descripcion", "").strip()

        marca = form_data.get("marca", "").strip()

        fecha_raw = form_data.get("fecha_ingreso", "").strip()

        calidad = form_data.get("calidad", "").strip()

        precio_raw = form_data.get("precio", "").strip()

        existencias_raw = form_data.get("existencias", "").strip()

        categoria = form_data.get("categoria", "").strip()

        imagen = request.FILES.get("imagen")



        errores = []
        if imagen:
            try:
                _validar_imagen_producto(imagen)
            except ValidationError as exc:
                errores.append(str(exc))

        if not nombre:

            errores.append("El nombre del producto es obligatorio.")

        if not marca:

            errores.append("La marca es obligatoria.")

        if not calidad:

            errores.append("Indica la calidad del producto.")

        if not categoria:

            errores.append("Selecciona una categoría.")



                          

        fecha_ingreso = None

        if fecha_raw:

            fecha_ingreso = parse_date(fecha_raw)

            if not fecha_ingreso:

                errores.append("La fecha de ingreso no tiene un formato válido (AAAA-MM-DD).")

        else:

            fecha_ingreso = timezone.localdate()



                

        precio = None

        if precio_raw:

            try:

                precio = Decimal(precio_raw)

                if precio < 0:

                    raise InvalidOperation

            except (InvalidOperation, TypeError):

                errores.append("Ingresa un precio válido.")

        else:

            errores.append("El precio es obligatorio.")



                     

        existencias = None

        if existencias_raw:

            try:

                existencias = int(existencias_raw)

                if existencias < 0:

                    raise ValueError

            except (TypeError, ValueError):

                errores.append("Las existencias deben ser un número entero positivo.")

        else:

            errores.append("Debes indicar existencias disponibles.")



        if errores:

            for error in errores:

                messages.error(request, error)

        else:

            if form_data.get("producto_id"):

                try:

                    p = Producto.objects.get(pk=form_data.get("producto_id"))

                except Producto.DoesNotExist:

                    messages.error(request, "Producto no encontrado.")

                    return redirect("dashboard_vendedor")

                                                

                if vendedor and p.vendedor_id and p.vendedor_id != vendedor.id:

                    messages.error(request, "No puedes editar productos de otro vendedor.")

                    return redirect("dashboard_vendedor")

                p.nombre = nombre

                p.descripcion = descripcion

                p.marca = marca

                p.fecha_ingreso = fecha_ingreso

                p.calidad = calidad

                p.precio = precio

                p.existencias = existencias

                p.categoria = categoria

                if vendedor and not p.vendedor_id:

                    p.vendedor = vendedor

                if imagen:

                    p.imagen = imagen

                p.save()

                messages.success(request, "Producto actualizado correctamente.")

                return redirect("dashboard_vendedor")

            else:

                Producto.objects.create(

                    vendedor=vendedor,

                    nombre=nombre,

                    descripcion=descripcion,

                    marca=marca,

                    fecha_ingreso=fecha_ingreso,

                    calidad=calidad,

                    precio=precio,

                    existencias=existencias,

                    categoria=categoria,

                    imagen=imagen,

                )

            messages.success(request, "Producto agregado correctamente.")

            return redirect("dashboard_vendedor")



    categorias_qs = Producto.objects.all()

    if vendedor:

        categorias_qs = categorias_qs.filter(vendedor=vendedor)

    categorias = categorias_qs.order_by("categoria").values_list("categoria", flat=True).distinct()



    contexto = {

        "productos": productos,

        "form_data": form_data,

        "categorias": categorias,

        "calidad_opciones": [

            "Nuevo",

            "Casi nuevo",

            "Coleccionista",

            "Usado",

        ],

        "today": timezone.localdate().isoformat(),

    }

    return render(request, "dashboards/dashboard_vendedor.html", contexto)





@login_required

@require_http_methods(["GET"])

def api_vendedor_resumen(request):

    """

    Devuelve métricas REALES para el vendedor actual:

    - ventas_hoy, ticket_promedio, serie úlltimos 7 días,

      ventas por categoría (top 5), totales.

    """

    vendedor = Vendedor.objects.filter(usuario=request.user).first()

    if not vendedor:

        return HttpResponseForbidden("No es vendedor")



    ventas_qs = Venta.objects.filter(vendedor=vendedor)



    total_ventas = float(ventas_qs.aggregate(s=Sum("total"))["s"] or 0)

    total_items = int(ventas_qs.aggregate(s=Sum("cantidad"))["s"] or 0)

    n_ventas = ventas_qs.count()

    ticket_prom = float((total_ventas / n_ventas) if n_ventas else 0)



                                        

    today = timezone.localdate()

    last7 = [today - timedelta(days=d) for d in range(6, -1, -1)]

    ventas_por_dia = ventas_qs.values("fecha_venta").annotate(total=Sum("total"))

    mapa_dias = {v["fecha_venta"]: float(v["total"] or 0) for v in ventas_por_dia}

    labels = [d.strftime("%a") for d in last7]                 

    data = [mapa_dias.get(d, 0) for d in last7]



                                

    por_categoria_qs = (

        ventas_qs

        .values(nombre=F("producto__categoria"))

        .annotate(total=Sum("total"))

        .order_by("-total")[:5]

    )

    por_categoria = [

        {"categoria": r["nombre"] or "Sin categoria", "total": float(r["total"] or 0)}

        for r in por_categoria_qs

    ]



    ventas_hoy = float(

        ventas_qs.filter(fecha_venta=today).aggregate(s=Sum("total"))["s"] or 0

    )



    return JsonResponse({

        "ventas_hoy": ventas_hoy,

        "ticket_promedio": ticket_prom,

        "tasa_conversion": 2.4,                                                      

        "labels": labels,

        "data": data,

        "por_categoria": por_categoria,

        "total_ventas": total_ventas,

        "total_items": total_items,

    })




@login_required

@require_http_methods(["GET"])

def api_vendedor_producto_detalle(request, pk):

    """

    Detalle de un producto para prefijar el formulario de edición del vendedor.

    """

    vendedor = Vendedor.objects.filter(usuario=request.user).first()

    try:

        p = Producto.objects.get(pk=pk)

    except Producto.DoesNotExist:

        return JsonResponse({"error": "not_found"}, status=404)

    if vendedor and p.vendedor_id and p.vendedor_id != vendedor.id:

        return JsonResponse({"error": "forbidden"}, status=403)



    data = {

        "id": p.id,

        "nombre": p.nombre or "",

        "marca": p.marca or "",

        "calidad": p.calidad or "",

        "categoria": p.categoria or "",

        "precio": float(p.precio or 0),

        "existencias": int(p.existencias or 0),

        "fecha": (p.fecha_ingreso.isoformat() if getattr(p, "fecha_ingreso", None) else ""),

        "descripcion": p.descripcion or "",

    }

    return JsonResponse(data)





                                                              

                                           

                                    

                                                        

                                                              



@login_required

def VistaAdministrador(request):

    """

    Renderiza el panel admin (si lo usas render server-side).

    En general, los datos ahora se consumen vía las APIs de abajo.

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo administradores")

    total_vendedores = Vendedor.objects.count()

    total_productos = Producto.objects.count()

    total_existencias = Producto.objects.aggregate(s=Sum("existencias"))["s"] or 0

    total_ventas = Venta.objects.aggregate(s=Sum("total"))["s"] or 0

    total_items = Venta.objects.aggregate(s=Sum("cantidad"))["s"] or 0



                                           

    ventas_por_vendedor = (

        Venta.objects

        .values("vendedor__usuario__username")

        .annotate(total=Sum("total"), cantidad=Sum("cantidad"))

        .order_by("-total")

    )

    labels_vendedores = [v["vendedor__usuario__username"] or "N/D" for v in ventas_por_vendedor]

    data_vendedores = [float(v["total"] or 0) for v in ventas_por_vendedor]



                                             

    hoy = timezone.localdate()

    hace_30 = hoy - timedelta(days=30)

    clientes_totales = Compra.objects.values_list("cliente", flat=True).distinct().count()

    clientes_activos_30 = (

        Compra.objects

        .filter(fecha_compra__gte=hace_30)

        .values_list("cliente", flat=True).distinct().count()

    )

    clientes_inactivos_30 = max(clientes_totales - clientes_activos_30, 0)



                              

    productos_bajo_stock = Producto.objects.filter(existencias__lte=5).order_by("existencias", "nombre")



    contexto = {

        "total_vendedores": total_vendedores,

        "total_productos": total_productos,

        "total_existencias": total_existencias,

        "total_ventas": total_ventas,

        "total_items_vendidos": total_items,

        "labels_vendedores": labels_vendedores,

        "data_vendedores": data_vendedores,

        "tabla_vendedores": ventas_por_vendedor,

        "productos_bajo_stock": productos_bajo_stock,

        "clientes_activos_30": clientes_activos_30,

        "clientes_inactivos_30": clientes_inactivos_30,

    }

                                                                  

    contexto["usuarios_iniciales"] = []

    contexto["stock_inicial"] = []

    return render(request, "dashboards/dashboard_administrador.html", contexto)





                                                                 



def _json_body(request):

    """Intenta decodificar el cuerpo JSON de la solicitud."""
    try:

        return json.loads(request.body.decode("utf-8") or "{}")

    except Exception:

        return {}



def _bad_request(message: str):

    """Devuelve una respuesta JSON de error con estado 400."""
    try:

        msg = str(message)

    except Exception:

        msg = "bad_request"

                                              

    try:

        print(f"api_admin_vendedores 400 -> {msg}")

    except Exception:

        pass

    return JsonResponse({"error": msg}, status=400)





                                                                 



@login_required

@require_http_methods(["GET"])

def api_admin_productos_bajo_stock(request):

    """

    Listado de productos.

    - Por defecto: solo críticos (existencias <= 5).

    - ?all=1 → devuelve TODOS los productos.

    - ?vendedor_id=XX → filtra por productos que ese vendedor vendió alguna vez.

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



    vendedor_id = request.GET.get("vendedor_id")

    ver_todos = request.GET.get("all") in ("1", "true", "True")



    qs = Producto.objects.all() if ver_todos else Producto.objects.filter(existencias__lte=5)



    if vendedor_id:

        ventas_sub = Venta.objects.filter(producto_id=OuterRef("pk"), vendedor_id=vendedor_id)

        qs = qs.annotate(vendido_por=Exists(ventas_sub)).filter(vendido_por=True)



    qs = qs.select_related("vendedor__usuario").order_by("vendedor__usuario__username", "existencias", "nombre")



    items = []

    for p in qs:

        vendedor_nombre = getattr(getattr(p.vendedor, "usuario", None), "username", None)

        imagen_url = None

        try:

            if p.imagen and hasattr(p.imagen, "url"):

                imagen_url = p.imagen.url

        except Exception:

            imagen_url = None

        items.append({

            "id": p.id,

            "nombre": p.nombre,

            "vendedor": vendedor_nombre or "N/D",

            "stock": int(p.existencias or 0),

            "existencias": int(p.existencias or 0),

            "critico": bool((p.existencias or 0) <= 5),

            "tipo": p.categoria or "-",

            "categoria": p.categoria or "-",

            "imagen": imagen_url,

            "descripcion": p.descripcion or "",

        })

    return JsonResponse({"items": items})





@login_required

@require_http_methods(["GET"])

def api_admin_producto_detalle(request, pk):

    """Devuelve la información detallada de un producto."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")

    try:

        p = Producto.objects.get(pk=pk)

    except Producto.DoesNotExist:

        return JsonResponse({"error": "not_found"}, status=404)

    imagen_url = None

    try:

        if p.imagen and hasattr(p.imagen, "url"):

            imagen_url = p.imagen.url

    except Exception:

        imagen_url = None

    vendedor_nombre = getattr(getattr(p.vendedor, "usuario", None), "username", None)

    return JsonResponse({

        "id": p.id,

        "nombre": p.nombre or "",

        "existencias": int(p.existencias or 0),

        "critico": bool((p.existencias or 0) <= 5),

        "descripcion": p.descripcion or "",

        "imagen": imagen_url,

        "vendedor": vendedor_nombre or "N/D",

        "categoria": p.categoria or "",

    })





@login_required

@require_http_methods(["POST"])

def api_admin_producto_update_full(request, pk):

    """Actualiza todos los campos editables de un producto."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")

    try:

        p = Producto.objects.get(pk=pk)

    except Producto.DoesNotExist:

        return JsonResponse({"error": "not_found"}, status=404)



    nombre = request.POST.get("nombre", "").strip()

    descripcion = request.POST.get("descripcion", "").strip()

    existencias_raw = request.POST.get("existencias", "").strip()

    categoria = request.POST.get("categoria", "").strip()

    imagen = request.FILES.get("imagen")



    if nombre:

        p.nombre = nombre

    p.descripcion = descripcion

    if categoria:

        p.categoria = categoria

    if existencias_raw != "":

        try:

            p.existencias = int(existencias_raw)

            if p.existencias < 0:

                return HttpResponseBadRequest("existencias debe ser entero >= 0")

        except ValueError:

            return HttpResponseBadRequest("existencias inválidas")

    if imagen is not None:
        try:
            _validar_imagen_producto(imagen)
        except ValidationError as exc:
            return HttpResponseBadRequest(str(exc))
        p.imagen = imagen

    p.save()

    return JsonResponse({"ok": True, "id": p.id})





@login_required

@require_http_methods(["DELETE"])

def api_admin_producto_delete(request, pk):

    """Elimina un producto desde el panel administrativo."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")

    try:

        p = Producto.objects.get(pk=pk)

    except Producto.DoesNotExist:

        return JsonResponse({"error": "not_found"}, status=404)

    p.delete()

    return JsonResponse({"ok": True})





@login_required

@require_http_methods(["DELETE", "POST"])

def api_vendedor_producto_delete(request, pk):

    """Permite a un vendedor eliminar un producto propio.



    Acepta DELETE (preferido) y POST (fallback para entornos que no permiten DELETE desde formularios).

    """

    try:

        vend = Vendedor.objects.get(usuario=request.user)

    except Vendedor.DoesNotExist:

        return HttpResponseForbidden("Solo vendedores")



    try:

        p = Producto.objects.get(pk=pk, vendedor=vend)

    except Producto.DoesNotExist:

        return JsonResponse({"error": "not_found"}, status=404)



    p.delete()

    return JsonResponse({"ok": True})





@login_required

@require_http_methods(["PUT", "PATCH"])

def api_admin_producto_update_stock(request, pk):

    """

    Actualiza existencias de un producto (botón Editar en tabla de stock admin).

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



    data = _json_body(request)

    nuevo_stock = data.get("existencias")

    if nuevo_stock is None:

        return HttpResponseBadRequest("Campo 'existencias' requerido")



    try:

        nuevo_stock = int(nuevo_stock)

        if nuevo_stock < 0:

            raise ValueError

    except ValueError:

        return HttpResponseBadRequest("existencias debe ser entero >= 0")



    try:

        p = Producto.objects.get(pk=pk)

    except Producto.DoesNotExist:

        return JsonResponse({"error": "not_found"}, status=404)



    p.existencias = nuevo_stock

    p.save(update_fields=["existencias"])

    return JsonResponse({"ok": True, "id": p.id, "nombre": p.nombre, "existencias": p.existencias})





                                                                 



@login_required

@require_http_methods(["GET"])

def api_admin_ventas_por_vendedor(request):

    """

    Devuelve ventas por vendedor.

    - Solo vendedores ACTIVOS (User.is_active=True).

    - Incluye vendedores sin ventas (0).

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



    labels, data, rows = [], [], []

    vendedor_id = request.GET.get("vendedor_id")

    vendedores_qs = Vendedor.objects.select_related("usuario").filter(usuario__is_active=True)

    if vendedor_id:

        try:

            vendedores_qs = vendedores_qs.filter(pk=int(vendedor_id))

        except (TypeError, ValueError):

            vendedores_qs = vendedores_qs.none()

    vendedores_activos = vendedores_qs.order_by("usuario__username")

    for v in vendedores_activos:

        tot = Venta.objects.filter(vendedor=v).aggregate(s=Sum("total"), c=Sum("cantidad"))

        total = float(tot["s"] or 0)

        cant = int(tot["c"] or 0)

        nombre = v.usuario.username if v.usuario else f"Vendedor {v.id}"

        labels.append(nombre)

        data.append(total)

        rows.append({"vendedor": nombre, "total": total, "cantidad": cant})



    return JsonResponse({"labels": labels, "data": data, "rows": rows})





@login_required

@require_http_methods(["GET"])

@login_required

@require_http_methods(["GET"])

def api_admin_clientes_actividad(request):

    """

    Clientes activos vs inactivos últimos 30 días (si lo sigues mostrando en el donut).

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



    hoy = timezone.localdate()

    hace_30 = hoy - timedelta(days=30)

    clientes_totales = Compra.objects.values_list("cliente", flat=True).distinct().count()

    activos = (

        Compra.objects

        .filter(fecha_compra__gte=hace_30)

        .values_list("cliente", flat=True).distinct().count()

    )

    inactivos = max(clientes_totales - activos, 0)

    return JsonResponse({"activos": activos, "inactivos": inactivos})





@login_required

@require_http_methods(["GET"])

def api_admin_vendedores_estado(request):

    """

    Cuenta Activos vs Inactivos de vendedores para el donut.



    Modos:

      - Por defecto (legacy): usa `User.is_active` (habilitado/deshabilitado).

      - Con `presence=1`: usa presencia reciente mediante `last_login` en una

        ventana de tiempo (segundos) indicada por `window` (30..900; default 180).

        En este modo, "Activo" = visto en la ventana; "Inactivo" = resto.

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



    vendedor_id = request.GET.get("vendedor_id")

    presence = request.GET.get("presence") in {"1", "true", "True", "yes"}

    base = User.objects.filter(groups__name="Vendedores")



    if presence:

        try:

            window_seconds = int(request.GET.get("window", 180))

        except (TypeError, ValueError):

            window_seconds = 180

        window_seconds = max(30, min(900, window_seconds))

        limite = timezone.now() - timedelta(seconds=window_seconds)



        if vendedor_id:

            try:

                v = Vendedor.objects.select_related("usuario").get(pk=int(vendedor_id))

                u = v.usuario

                online = int(bool(u and u.is_active and u.last_login and u.last_login >= limite))

                offline = 1 - online

                return JsonResponse({"activos": online, "inactivos": offline, "scoped": True, "presence": True, "window": window_seconds})

            except (Vendedor.DoesNotExist, ValueError, TypeError):

                return JsonResponse({"activos": 0, "inactivos": 0, "scoped": True, "presence": True, "window": window_seconds})



                                                    

        online = base.filter(is_active=True, last_login__gte=limite).count()

        total = base.count()

        offline = max(0, total - online)

        return JsonResponse({"activos": online, "inactivos": offline, "scoped": False, "presence": True, "window": window_seconds})



                                                

    if vendedor_id:

        try:

            v = Vendedor.objects.select_related("usuario").get(pk=int(vendedor_id))

            activos = 1 if v.usuario and v.usuario.is_active else 0

            inactivos = 0 if activos == 1 else 1

            return JsonResponse({"activos": activos, "inactivos": inactivos, "scoped": True, "presence": False})

        except (Vendedor.DoesNotExist, ValueError, TypeError):

            return JsonResponse({"activos": 0, "inactivos": 0, "scoped": True, "presence": False})



    activos = base.filter(is_active=True).count()

    inactivos = base.filter(is_active=False).count()

    return JsonResponse({"activos": activos, "inactivos": inactivos, "scoped": False, "presence": False})





@login_required

@require_http_methods(["GET"])

def api_admin_usuarios_online(request):

    """Devuelve los IDs de usuarios activos en una ventana de tiempo reciente.



    Query params:

      - window: segundos de ventana (30..900, por defecto 180)

    Respuesta: { active_ids: [int, ...], window_seconds: int, generated_at: iso }

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



    try:

        window_seconds = int(request.GET.get("window", 180))

    except (TypeError, ValueError):

        window_seconds = 180

    window_seconds = max(30, min(900, window_seconds))



    limite = timezone.now() - timedelta(seconds=window_seconds)

    activos = list(

        User.objects.filter(is_active=True, last_login__gte=limite)

        .values_list("id", flat=True)

    )

    return JsonResponse({

        "active_ids": [int(pk) for pk in activos],

        "window_seconds": window_seconds,

        "generated_at": timezone.now().isoformat(),

    })





@login_required

@require_http_methods(["GET"])

def api_admin_top_productos_linea(request):

    """

    Serie por día de los TOP N productos (por total vendido) en los

    últimos `days` días. Opcionalmente filtrado por `vendedor_id`.

    Params:

      - vendedor_id (opcional)

      - days (opcional, default 30)

      - top_n (opcional, default 3)

    Respuesta:

      { labels: ["YYYY-MM-DD", ...], datasets: [{ label, data: [..] , product_id } ...] }

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



    vendedor_id = request.GET.get("vendedor_id")

    try:

        days = int(request.GET.get("days", 30))

        if days <= 0:

            days = 30

    except (TypeError, ValueError):

        days = 30

    try:

        top_n = int(request.GET.get("top_n", 3))

        if top_n <= 0:

            top_n = 3

    except (TypeError, ValueError):

        top_n = 3



    hoy = timezone.localdate()

    desde = hoy - timedelta(days=days - 1)



    ventas = Venta.objects.filter(fecha_venta__gte=desde, fecha_venta__lte=hoy)

    if vendedor_id:

        try:

            ventas = ventas.filter(vendedor_id=int(vendedor_id))

        except (TypeError, ValueError):

            ventas = ventas.none()



                                           

    top = (

        ventas

        .values("producto_id", "producto__nombre")

        .annotate(total=Sum("total"))

        .order_by("-total")[:top_n]

    )

    top_ids = [r["producto_id"] for r in top]

    top_names = {r["producto_id"]: (r["producto__nombre"] or f"Producto {r['producto_id']}") for r in top}



                             

    dias = [desde + timedelta(days=i) for i in range(days)]

    labels = [d.isoformat() for d in dias]



                                       

    datasets = []

    for pid in top_ids:

        serie = ventas.filter(producto_id=pid).values("fecha_venta").annotate(t=Sum("total"))

        mapa = {row["fecha_venta"].isoformat(): float(row["t"] or 0) for row in serie}

        data = [mapa.get(lbl, 0.0) for lbl in labels]

        datasets.append({

            "label": top_names.get(pid, f"Producto {pid}"),

            "data": data,

            "product_id": pid,

        })



    return JsonResponse({"labels": labels, "datasets": datasets})





                                                                



@login_required

@require_http_methods(["GET", "POST", "PUT", "PATCH", "DELETE"])

def api_admin_vendedores(request):

    """

    CRUD básico de usuarios-vendedores.

    GET:   lista (con filtros q y estado)

    POST:  crear (username, email?, password) -> añade al grupo 'Vendedores' y crea Vendedor

    PUT:   editar (id + campos)

    DEL:   toggle is_active (activar/desactivar)

    """

    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")



                                  

    if request.method == "GET":

        estado = request.GET.get("estado")                                      

        q = (request.GET.get("q") or "").strip()



        usuarios = User.objects.all().order_by("id")

        if estado == "activo":

            usuarios = usuarios.filter(is_active=True)

        elif estado == "inactivo":

            usuarios = usuarios.filter(is_active=False)



        if q:

            usuarios = usuarios.filter(username__icontains=q) | usuarios.filter(email__icontains=q)



        vend_map = {u_id: v_id for v_id, u_id in Vendedor.objects.values_list("id", "usuario_id")}

        items = []

        for u in usuarios:

            es_vend = u.id in vend_map

            es_admin = bool(u.is_staff or u.is_superuser)

            rol = "Administrador" if es_admin else ("Vendedor" if es_vend else "Usuario")

            items.append({

                "id": u.id,

                "username": u.username,

                "email": u.email or "",

                "date_joined": u.date_joined.strftime("%Y-%m-%d"),

                "last_login": (u.last_login.strftime("%Y-%m-%d %H:%M") if u.last_login else "-"),

                "is_active": u.is_active,

                "es_vendedor": es_vend,

                "es_admin": es_admin,

                "is_self": u.id == request.user.id,

                "role": rol,

                "vendedor_id": vend_map.get(u.id),

                "password_hash": u.password if (request.user.is_superuser or request.user.is_staff) else "****",

            })

        return JsonResponse({"items": items})



    data = _json_body(request)



                                 

    if request.method == "POST":

        try:

            username = (data.get("username") or "").strip()

            email = (data.get("email") or "").strip()

            password = (data.get("password") or "").strip()

            if not username or not password:

                return _bad_request("username y password son obligatorios")

            if len(username) < 3:

                return _bad_request("username debe tener al menos 3 caracteres")

            if User.objects.filter(username=username).exists():

                return _bad_request("username ya existe")

                                                                

            if email:

                try:

                    validate_email(email)

                except ValidationError:

                    return _bad_request("email invalido")

                if User.objects.filter(email__iexact=email).exists():

                    return _bad_request("email ya esta en uso")

                                                                      

            try:

                validate_password(password)

            except ValidationError as e:

                return _bad_request(" ".join([str(m) for m in e.messages]))



            u = User.objects.create_user(username=username, email=email, password=password)

                                                            

            es_vendedor = bool(data.get("es_vendedor"))

            es_admin = bool(data.get("es_admin"))

            u.is_active = True

            if es_admin:

                u.is_staff = True

            u.save()

            if es_vendedor:

                grupo, _ = Group.objects.get_or_create(name="Vendedores")

                u.groups.add(grupo)

                try:
                    Vendedor.objects.get_or_create(usuario=u)
                except Exception:
                    try:
                        with connection.cursor() as cur:
                            cur.execute(
                                "INSERT INTO core_vendedor (usuario_id, telefono, direccion, fecha_ingreso, umbral_critico) "
                                "VALUES (%s, %s, %s, CURRENT_DATE, %s)",
                                [u.id, "", "", 5],
                            )
                    except Exception as e2:
                        return _bad_request(f"server: {e2}")

                                                                                

                try:

                    if email:

                        asunto = "Bienvenido como vendedor a EpicAnimes!"

                        panel_url = request.build_absolute_uri(reverse("dashboard_vendedor"))

                        cuerpo = (

                            f"Hola {username},\n\n"

                            "Bienvenido a EpicAnimes! Tu cuenta de vendedor ha sido creada y aprobada.\n"

                            "Desde ahora puedes ingresar a tu panel para publicar y administrar productos, revisar ventas y gestionar tu catálogo.\n\n"

                            f"Panel de vendedor: {panel_url}\n\n"

                            "Si tienes dudas o necesitas ayuda, responde a este correo y nuestro equipo te apoyará.\n\n"

                            "¡éxito en tus ventas!"

                        )

                        send_mail(asunto, cuerpo, settings.DEFAULT_FROM_EMAIL, [email], fail_silently=False)

                except Exception:

                                                             

                    pass

            return JsonResponse({"ok": True, "id": u.id})

        except Exception as e:

            return _bad_request(f"server: {e}")



                                  

    if request.method in ("PUT", "PATCH"):

        uid = data.get("id")

        if not uid:

            return _bad_request("id requerido")

        try:

            u = User.objects.get(id=uid)

        except User.DoesNotExist:

            return _bad_request("usuario no existe")



                                                                                       

        if data.get("reset_password"):

            tmp = User.objects.make_random_password(length=10)

            u.set_password(tmp)

            u.save()

            return JsonResponse({"ok": True, "temp_password": tmp})



        username = data.get("username")

        email = data.get("email")

        password = data.get("password")

        is_active = data.get("is_active")

        es_vendedor = data.get("es_vendedor", True)



        if username:

            if User.objects.exclude(id=u.id).filter(username=username).exists():

                return HttpResponseBadRequest("username ya en uso")

            u.username = username

        if email is not None:

            email_s = (email or "").strip()

            if email_s:

                try:

                    validate_email(email_s)

                except ValidationError:

                    return _bad_request("email invalido")

                if User.objects.exclude(id=u.id).filter(email__iexact=email_s).exists():

                    return _bad_request("email ya esta en uso")

            u.email = email_s

        if isinstance(is_active, bool):

            u.is_active = is_active

        if password:

            try:

                validate_password(password, user=u)

            except ValidationError as e:

                return _bad_request(" ".join([str(m) for m in e.messages]))

            u.set_password(password)

        u.save()



        grupo, _ = Group.objects.get_or_create(name="Vendedores")

        if es_vendedor:

            try:
                Vendedor.objects.get_or_create(usuario=u)
            except Exception:
                try:
                    with connection.cursor() as cur:
                        cur.execute(
                            "INSERT INTO core_vendedor (usuario_id, telefono, direccion, fecha_ingreso, umbral_critico) "
                            "VALUES (%s, %s, %s, CURRENT_DATE, %s)",
                            [u.id, "", "", 5],
                        )
                except Exception as e2:
                    return _bad_request(f"server: {e2}")

            u.groups.add(grupo)

        else:

            Vendedor.objects.filter(usuario=u).delete()

            u.groups.remove(grupo)



        return JsonResponse({"ok": True})



                                                

    if request.method == "DELETE":

        uid = data.get("id")

        if not uid:

            return _bad_request("id requerido")

        try:

            u = User.objects.get(id=uid)

        except User.DoesNotExist:

            return _bad_request("usuario no existe")



        if data.get("eliminar"):

            if u.id == request.user.id:

                return _bad_request("No puedes eliminar tu propio usuario.")

            if u.is_superuser and not request.user.is_superuser:

                return HttpResponseForbidden("No puedes eliminar este usuario.")

            Vendedor.objects.filter(usuario=u).delete()

            u.delete()

            return JsonResponse({"ok": True, "eliminado": True})



        u.is_active = not u.is_active

        u.save()

        return JsonResponse({"ok": True, "is_active": u.is_active})



    return HttpResponseNotAllowed(["GET", "POST", "PUT", "PATCH", "DELETE"])





                                                              

                                                   

                                                              



@login_required

def redireccion_usuario(request):

    """Redirige al usuario al panel correspondiente según su rol."""
    user = request.user

    if user.is_superuser or user.is_staff:

        return redirect("dashboard_administrador")

    if user.groups.filter(name="Vendedores").exists():

        return redirect("dashboard_vendedor")

    return redirect("index")



                                                                         



@login_required

@require_http_methods(["GET", "PUT", "PATCH"])

def api_admin_postulaciones(request):

    """Entrega las postulaciones de vendedores en formato JSON."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")

    if request.method == "GET":

        qs = PostulacionVendedor.objects.all().order_by("-fecha_envio")

        q = (request.GET.get("q") or "").strip().lower()

        estado = (request.GET.get("estado") or "").strip().lower()

        desde_s = request.GET.get("from")

        hasta_s = request.GET.get("to")

        if estado in {"nuevo", "contactado", "archivado"}:

            qs = qs.filter(estado=estado)

        if q:

            qs = qs.filter(Q(nombre__icontains=q) | Q(email__icontains=q) | Q(mensaje__icontains=q))

        try:

            d = parse_date(desde_s) if desde_s else None

            if d:

                qs = qs.filter(fecha_envio__date__gte=d)

        except Exception:

            pass

        try:

            h = parse_date(hasta_s) if hasta_s else None

            if h:

                qs = qs.filter(fecha_envio__date__lte=h)

        except Exception:

            pass

        items = [

            {

                "id": p.id,

                "nombre": p.nombre,

                "email": p.email,

                "telefono": p.telefono or "",

                "tienda": p.tienda or "",

                "instagram": p.instagram or "",

                "mensaje": p.mensaje or "",

                "notas": p.notas or "",

                "fecha_envio": p.fecha_envio.strftime("%Y-%m-%d %H:%M"),

                "estado": p.estado,

            }

            for p in qs

        ]

        return JsonResponse({"items": items})



    data = _json_body(request)

    pid = data.get("id")

    nuevo_estado = (data.get("estado") or "").strip().lower()

    notas = data.get("notas")

    if not pid and notas is None:

        return HttpResponseBadRequest("id requerido")

    try:

        p = PostulacionVendedor.objects.get(pk=int(pid))

    except (PostulacionVendedor.DoesNotExist, ValueError, TypeError):

        return JsonResponse({"error": "not_found"}, status=404)

    updates = []

    if nuevo_estado in {"nuevo", "contactado", "archivado"}:

        p.estado = nuevo_estado

        updates.append("estado")

    if isinstance(notas, str):

        p.notas = notas

        updates.append("notas")

    if not updates:

        return HttpResponseBadRequest("sin cambios")

    p.save(update_fields=updates)

    return JsonResponse({"ok": True, "id": p.id, "estado": p.estado})





                                                                



@login_required

@require_http_methods(["GET"])

def export_admin_postulaciones_csv(request):

    """Genera un CSV con las postulaciones recibidas."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")

    import csv

    resp = HttpResponse(content_type='text/csv; charset=utf-8')

    resp["Content-Disposition"] = 'attachment; filename="postulaciones.csv"'

    writer = csv.writer(resp)

    writer.writerow(["id","nombre","email","telefono","tienda","instagram","mensaje","notas","fecha_envio","estado"])

    qs = PostulacionVendedor.objects.all().order_by("-fecha_envio")

    q = (request.GET.get("q") or "").strip().lower()

    estado = (request.GET.get("estado") or "").strip().lower()

    desde_s = request.GET.get("from")

    hasta_s = request.GET.get("to")

    if estado in {"nuevo","contactado","archivado"}: qs = qs.filter(estado=estado)

    if q: qs = qs.filter(Q(nombre__icontains=q)|Q(email__icontains=q)|Q(mensaje__icontains=q))

    try:

        d=parse_date(desde_s) if desde_s else None

        if d: qs=qs.filter(fecha_envio__date__gte=d)

    except Exception: pass

    try:

        h=parse_date(hasta_s) if hasta_s else None

        if h: qs=qs.filter(fecha_envio__date__lte=h)

    except Exception: pass

    for p in qs:

        writer.writerow([p.id, p.nombre, p.email, p.telefono or '', p.tienda or '', p.instagram or '', (p.mensaje or '').replace('\n',' ').strip(), (p.notas or '').replace('\n',' ').strip(), p.fecha_envio.isoformat(sep=' '), p.estado])

    return resp





@login_required

@require_http_methods(["GET"])

def export_admin_ventas_csv(request):

    """Exporta las ventas en formato CSV para administradores."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo admin")

    import csv

    desde_s = request.GET.get('from') or request.GET.get('start')

    hasta_s = request.GET.get('to') or request.GET.get('end')

    days = request.GET.get('days')

    qs = Venta.objects.select_related('vendedor__usuario','producto')

    if desde_s:

        try:

            d = parse_date(desde_s)

            if d: qs = qs.filter(fecha_venta__gte=d)

        except Exception: pass

    if hasta_s:

        try:

            h = parse_date(hasta_s)

            if h: qs = qs.filter(fecha_venta__lte=h)

        except Exception: pass

    if not desde_s and not hasta_s and days:

        try:

            n = int(days)

            n = max(1, min(365, n))

            hoy = timezone.localdate()

            qs = qs.filter(fecha_venta__gte=hoy - timedelta(days=n))

        except Exception: pass

    resp = HttpResponse(content_type='text/csv; charset=utf-8')

    resp["Content-Disposition"] = 'attachment; filename="ventas.csv"'

    writer = csv.writer(resp)

    writer.writerow(["fecha","vendedor","producto","cantidad","total"])

    for v in qs.order_by('fecha_venta','vendedor_id'):

        writer.writerow([v.fecha_venta.isoformat(), getattr(getattr(v.vendedor,'usuario',None),'username',''), v.producto.nombre, v.cantidad, f"{v.total}"])

    return resp





@login_required

@require_http_methods(["GET"])

def export_vendedor_inventario_csv(request):

                   

    """Construye un CSV con el inventario del vendedor."""
    if not request.user.groups.filter(name="Vendedores").exists():

        return HttpResponseForbidden("Solo vendedores")

    try:

        vend = Vendedor.objects.get(usuario=request.user)

    except Vendedor.DoesNotExist:

        return HttpResponseForbidden("Perfil vendedor requerido")

    import csv

    resp = HttpResponse(content_type='text/csv; charset=utf-8')

    resp["Content-Disposition"] = 'attachment; filename="inventario.csv"'

    writer = csv.writer(resp)

    writer.writerow(["nombre","marca","calidad","categoria","precio","existencias","fecha_ingreso","descripcion"])

    for p in Producto.objects.filter(vendedor=vend).order_by('nombre'):

        writer.writerow([p.nombre, p.marca, p.calidad, p.categoria, f"{p.precio}", int(p.existencias or 0), (p.fecha_ingreso.isoformat() if p.fecha_ingreso else ''), (p.descripcion or '').replace('\n',' ')])

    return resp





@login_required

@require_http_methods(["GET"])

def export_vendedor_ventas_csv(request):

    """Exporta las ventas del vendedor en CSV."""
    if not request.user.groups.filter(name="Vendedores").exists():

        return HttpResponseForbidden("Solo vendedores")

    try:

        vend = Vendedor.objects.get(usuario=request.user)

    except Vendedor.DoesNotExist:

        return HttpResponseForbidden("Perfil vendedor requerido")

    desde_s = request.GET.get('from')

    hasta_s = request.GET.get('to')

    days = request.GET.get('days')

    qs = Venta.objects.filter(vendedor=vend).select_related('producto')

    if desde_s:

        d = parse_date(desde_s)

        if d: qs = qs.filter(fecha_venta__gte=d)

    if hasta_s:

        h = parse_date(hasta_s)

        if h: qs = qs.filter(fecha_venta__lte=h)

    if not desde_s and not hasta_s and days:

        try:

            n = int(days)

            n = max(1, min(365, n))

            hoy = timezone.localdate()

            qs = qs.filter(fecha_venta__gte=hoy - timedelta(days=n))

        except Exception:

            pass

    import csv

    resp = HttpResponse(content_type='text/csv; charset=utf-8')

    resp["Content-Disposition"] = 'attachment; filename="ventas_vendedor.csv"'

    w = csv.writer(resp)

    w.writerow(["fecha","producto","cantidad","total"])

    for v in qs.order_by('fecha_venta'):

        w.writerow([v.fecha_venta.isoformat(), v.producto.nombre, v.cantidad, f"{v.total}"])

    return resp





@login_required

@require_http_methods(["POST"])

def api_vendedor_importar(request):

                   

    """Permite cargar productos en lote para el vendedor autenticado."""
    if not request.user.groups.filter(name="Vendedores").exists():

        return HttpResponseForbidden("Solo vendedores")

    try:

        vend = Vendedor.objects.get(usuario=request.user)

    except Vendedor.DoesNotExist:

        return HttpResponseForbidden("Perfil vendedor requerido")

    f = request.FILES.get('file') or request.FILES.get('csv')

    if not f:

        return HttpResponseBadRequest("Archivo 'file' CSV requerido")

    import csv, io

    try:

        data = f.read().decode('utf-8', errors='ignore')

        sample = data[:1024]

        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")

    except Exception:

        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(data), dialect=dialect)

    created = 0

    for row in reader:

        nombre = (row.get('nombre') or row.get('Nombre') or '').strip()

        if not nombre:

            continue

        marca = (row.get('marca') or row.get('Marca') or '').strip()

        calidad = (row.get('calidad') or row.get('Calidad') or '').strip()

        categoria = (row.get('categoria') or row.get('Categoría') or row.get('Categoria') or '').strip()

        try:

            precio = Decimal(str(row.get('precio') or row.get('Precio') or '0')).quantize(Decimal('0.01'))

        except Exception:

            precio = Decimal('0.00')

        try:

            exist = int((row.get('existencias') or row.get('stock') or row.get('Existencias') or '0').strip() or 0)

        except Exception:

            exist = 0

        fecha = parse_date((row.get('fecha_ingreso') or row.get('Fecha') or '').strip() or '') or timezone.localdate()

        desc = (row.get('descripcion') or row.get('Descripción') or '').strip()

        Producto.objects.create(

            vendedor=vend,

            nombre=nombre,

            marca=marca,

            calidad=calidad,

            categoria=categoria,

            precio=precio,

            existencias=exist,

            fecha_ingreso=fecha,

            descripcion=desc,

        )

        created += 1

    return JsonResponse({"ok": True, "creados": created})



                       

                       

                       

@login_required

@require_http_methods(["GET"])

def export_admin_postulaciones_xlsx(request):

    """Genera un archivo XLSX de postulaciones."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo administradores")

    try:

        from openpyxl import Workbook

    except Exception:

        return HttpResponseBadRequest("Falta dependencia 'openpyxl'")

    from io import BytesIO

    wb = Workbook()

    ws = wb.active

    ws.title = "Postulaciones"

    ws.append(["ID","Nombre","Email","Teléfono","Tienda","Instagram/Web","Mensaje","Fecha","Estado","Notas"])

    q = (request.GET.get("q") or "").strip().lower()

    estado = (request.GET.get("estado") or "").strip().lower()

    desde_s = request.GET.get("from")

    hasta_s = request.GET.get("to")

    qs = PostulacionVendedor.objects.all()

    if estado in {"nuevo","contactado","archivado"}:

        qs = qs.filter(estado=estado)

    if q:

        qs = qs.filter(Q(nombre__icontains=q) | Q(email__icontains=q) | Q(mensaje__icontains=q))

    try:

        d = parse_date(desde_s) if desde_s else None

        if d:

            qs = qs.filter(fecha_envio__date__gte=d)

    except Exception:

        pass

    try:

        h = parse_date(hasta_s) if hasta_s else None

        if h:

            qs = qs.filter(fecha_envio__date__lte=h)

    except Exception:

        pass

    for p in qs.order_by("-fecha_envio"):

        ws.append([

            p.id,

            p.nombre,

            p.email,

            p.telefono or "",

            p.tienda or "",

            p.instagram or "",

            (p.mensaje or "")[:500],

            p.fecha_envio.isoformat(sep=" "),

            p.estado,

            (p.notas or "")[:500],

        ])

    bio = BytesIO(); wb.save(bio); bio.seek(0)

    resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    resp["Content-Disposition"] = 'attachment; filename="postulaciones.xlsx"'

    return resp





@login_required

@require_http_methods(["GET"])

def export_admin_ventas_xlsx(request):

    """Exporta las ventas en formato XLSX."""
    if not (request.user.is_staff or request.user.is_superuser):

        return HttpResponseForbidden("Solo administradores")

    try:

        from openpyxl import Workbook

    except Exception:

        return HttpResponseBadRequest("Falta dependencia 'openpyxl'")

    desde_s = request.GET.get('from') or request.GET.get('start')

    hasta_s = request.GET.get('to') or request.GET.get('end')

    days = request.GET.get('days')

    qs = Venta.objects.select_related('vendedor__usuario','producto')

    if desde_s:

        try:

            d = parse_date(desde_s)

            if d: qs = qs.filter(fecha_venta__gte=d)

        except Exception:

            pass

    if hasta_s:

        try:

            h = parse_date(hasta_s)

            if h: qs = qs.filter(fecha_venta__lte=h)

        except Exception:

            pass

    if not desde_s and not hasta_s and days:

        try:

            n = int(days)

            n = max(1, min(365, n))

            hoy = timezone.localdate()

            qs = qs.filter(fecha_venta__gte=hoy - timedelta(days=n))

        except Exception:

            pass

    from io import BytesIO

    wb = Workbook(); ws = wb.active; ws.title = "Ventas"

    ws.append(["Fecha","Vendedor","Producto","Cantidad","Total"])

    for v in qs.order_by('fecha_venta','vendedor_id'):

        ws.append([

            v.fecha_venta.isoformat(),

            getattr(getattr(v.vendedor,'usuario',None),'username',''),

            v.producto.nombre,

            v.cantidad,

            float(v.total),

        ])

    bio = BytesIO(); wb.save(bio); bio.seek(0)

    resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    resp["Content-Disposition"] = 'attachment; filename="ventas.xlsx"'

    return resp





@login_required

@require_http_methods(["GET"])

def export_vendedor_inventario_xlsx(request):

    """Genera un XLSX del inventario del vendedor."""
    if not request.user.groups.filter(name="Vendedores").exists():

        return HttpResponseForbidden("Solo vendedores")

    try:

        vend = Vendedor.objects.get(usuario=request.user)

    except Vendedor.DoesNotExist:

        return HttpResponseForbidden("Perfil vendedor requerido")

    try:

        from openpyxl import Workbook

    except Exception:

        return HttpResponseBadRequest("Falta dependencia 'openpyxl'")

    from io import BytesIO

    wb = Workbook(); ws = wb.active; ws.title = "Inventario"

    ws.append(["Nombre","Marca","Calidad","Categoría","Precio","Existencias","Fecha ingreso","Descripción"])

    for p in Producto.objects.filter(vendedor=vend).order_by('nombre'):

        ws.append([

            p.nombre,

            p.marca,

            p.calidad,

            p.categoria,

            float(p.precio),

            int(p.existencias or 0),

            (p.fecha_ingreso.isoformat() if p.fecha_ingreso else ''),

            (p.descripcion or ''),

        ])

    bio = BytesIO(); wb.save(bio); bio.seek(0)

    resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    resp["Content-Disposition"] = 'attachment; filename="inventario.xlsx"'

    return resp





@login_required

@require_http_methods(["GET"])

def export_vendedor_ventas_xlsx(request):

    """Exporta las ventas del vendedor en XLSX."""
    if not request.user.groups.filter(name="Vendedores").exists():

        return HttpResponseForbidden("Solo vendedores")

    try:

        vend = Vendedor.objects.get(usuario=request.user)

    except Vendedor.DoesNotExist:

        return HttpResponseForbidden("Perfil vendedor requerido")

    try:

        from openpyxl import Workbook

    except Exception:

        return HttpResponseBadRequest("Falta dependencia 'openpyxl'")

    desde_s = request.GET.get('from')

    hasta_s = request.GET.get('to')

    days = request.GET.get('days')

    qs = Venta.objects.filter(vendedor=vend).select_related('producto')

    if desde_s:

        d = parse_date(desde_s)

        if d: qs = qs.filter(fecha_venta__gte=d)

    if hasta_s:

        h = parse_date(hasta_s)

        if h: qs = qs.filter(fecha_venta__lte=h)

    if not desde_s and not hasta_s and days:

        try:

            n = int(days)

            n = max(1, min(365, n))

            hoy = timezone.localdate()

            qs = qs.filter(fecha_venta__gte=hoy - timedelta(days=n))

        except Exception:

            pass

    from io import BytesIO

    wb = Workbook(); ws = wb.active; ws.title = "Ventas"

    ws.append(["Fecha","Producto","Cantidad","Total"])

    for v in qs.order_by('fecha_venta'):

        ws.append([v.fecha_venta.isoformat(), v.producto.nombre, v.cantidad, float(v.total)])

    bio = BytesIO(); wb.save(bio); bio.seek(0)

    resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    resp["Content-Disposition"] = 'attachment; filename="ventas_vendedor.xlsx"'

    return resp





                            

                        

                            

@login_required
@require_http_methods(["POST"])
def api_vendedor_importar_excel(request):
    """Procesa archivos Excel para crear o actualizar productos."""
    if not request.user.groups.filter(name="Vendedores").exists():
        return HttpResponseForbidden("Solo vendedores")
    try:
        vend = Vendedor.objects.get(usuario=request.user)
    except Vendedor.DoesNotExist:
        return HttpResponseForbidden("Perfil vendedor requerido")
    f = request.FILES.get('file')
    if not f:
        return HttpResponseBadRequest("Archivo Excel requerido")
    try:
        from openpyxl import load_workbook
        from io import BytesIO
        content = f.read()
        wb = load_workbook(filename=BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return JsonResponse({"ok": True, "creados": 0})
        def _norm(s):
            if not s:
                return ''
            import unicodedata
            s = str(s)
            s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
            return s.strip()
        headers = [(_norm(h) or '').lower().replace(' ', '_') for h in (rows[0] or [])]
        created = 0
        for r in rows[1:]:
            row = {}
            for idx, val in enumerate(r or []):
                key = headers[idx] if idx < len(headers) else f'col_{idx}'
                row[key] = val if val is not None else ''
            nombre = _norm(row.get('nombre') or row.get('name'))
            if not nombre:
                continue
            marca = _norm(row.get('marca'))
            calidad = _norm(row.get('calidad'))
            categoria = _norm(row.get('categoria') or row.get('categoria_'))
            try:
                precio = Decimal(str(row.get('precio') or '0')).quantize(Decimal('0.01'))
            except Exception:
                precio = Decimal('0.00')
            try:
                exist = int(str(row.get('existencias') or row.get('stock') or '0'))
            except Exception:
                exist = 0
            fecha_raw = str(row.get('fecha_ingreso') or '')
            fecha = parse_date(fecha_raw) or timezone.localdate()
            desc = str(row.get('descripcion') or '')
            Producto.objects.create(
                vendedor=vend,
                nombre=nombre,
                marca=marca,
                calidad=calidad,
                categoria=categoria,
                precio=precio,
                existencias=exist,
                fecha_ingreso=fecha,
                descripcion=desc,
            )
            created += 1
        return JsonResponse({"ok": True, "creados": created})
    except Exception:
        return HttpResponseBadRequest("No se pudo procesar el Excel")




class CoreLoginView(LoginView):

    """Renderiza el formulario de inicio de sesión con verificación OTP."""
    template_name = 'registration/login.html'

    authentication_form = TwoFactorLoginForm





# GRÁFICO DE BARRA HORIZONTAL EN DASHBOARD ADMINISTRADOR
@login_required
@require_http_methods(["GET"])
def api_admin_ventas_por_usuario(request):
    """
    Ranking (barra horizontal) de ventas por usuario en un rango de días.

    - Aggrega Compras por usuario: sum(valor_producto * cantidad)
    - Permite `top` resultados (3..50)
    - Opcional `presence=1&window=180` para devolver conteos por rol
      basados en ventana de presencia (sino, 30 días por último acceso).

    Respuesta:
      { labels: [username], data: [total_CLP], counts: { usuarios: {activos,inactivos,suspendidos}, vendedores: {...}, administradores: {...} } }
    """

    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Solo admin")

                 
    try:
        days = int(request.GET.get("days", 30))
    except (TypeError, ValueError):
        days = 30
    days = max(1, min(days, 365))

    try:
        top_n = int(request.GET.get("top", 10))
    except (TypeError, ValueError):
        top_n = 10
    top_n = max(3, min(top_n, 50))

    hoy = timezone.localdate()
    desde = hoy - timedelta(days=days - 1)

                                              
    compras = (
        Compra.objects
        .filter(fecha_compra__gte=desde, fecha_compra__lte=hoy, usuario__isnull=False)
        .values("usuario_id", "usuario__username")
        .annotate(
            total=Sum(F("valor_producto") * F("cantidad")),
            ordenes=Count("id"),
        )
        .order_by("-total")
    )[:top_n]

    labels = [row["usuario__username"] or f"Usuario {row['usuario_id']}" for row in compras]
    data = [float(row["total"] or 0) for row in compras]

                                                         
    presence = (request.GET.get("presence") == "1")
    if presence:
        try:
            window_seconds = int(request.GET.get("window", 180))
        except (TypeError, ValueError):
            window_seconds = 180
        window_seconds = max(30, min(window_seconds, 3600))
        limite_dt = timezone.now() - timedelta(seconds=window_seconds)
    else:
        limite_dt = timezone.now() - timedelta(days=30)

    base = User.objects.all()
    admins = base.filter(Q(is_staff=True) | Q(is_superuser=True))
    vendedores = base.filter(groups__name="Vendedores", is_staff=False, is_superuser=False)
    usuarios = base.filter(is_staff=False, is_superuser=False).exclude(groups__name="Vendedores")

    def counts_for(qs):
        suspendidos = qs.filter(is_active=False).count()
        activos = qs.filter(is_active=True, last_login__gte=limite_dt).count()
        total_activos = qs.filter(is_active=True).count()
        inactivos = max(0, total_activos - activos)
        return {"activos": int(activos), "inactivos": int(inactivos), "suspendidos": int(suspendidos)}

    counts = {
        "usuarios": counts_for(usuarios),
        "vendedores": counts_for(vendedores),
        "administradores": counts_for(admins),
    }

                                                
    agg_global = (
        Compra.objects
        .filter(fecha_compra__gte=desde, fecha_compra__lte=hoy)
        .aggregate(
            total_ventas=Sum(F("valor_producto") * F("cantidad")),
            ordenes=Count("id"),
            compradores=Count("usuario", distinct=True),
        )
    )
    total_ventas = float(agg_global.get("total_ventas") or 0)
    ordenes_total = int(agg_global.get("ordenes") or 0)
    compradores_distintos = int(agg_global.get("compradores") or 0)
    avg_ticket = float(total_ventas / ordenes_total) if ordenes_total else 0.0

                                            
    top3 = [
        {
            "name": row["usuario__username"] or f"Usuario {row['usuario_id']}",
            "total": float(row["total"] or 0),
            "ordenes": int(row["ordenes"] or 0),
        }
        for row in list(compras)[:3]
    ]

    return JsonResponse({
        "labels": labels,
        "data": data,
        "counts": counts,
        "summary": {
            "total_sold": total_ventas,
            "orders": ordenes_total,
            "buyers_considered": len(labels),
            "buyers_distinct": compradores_distintos,
            "avg_ticket": avg_ticket,
            "top_buyers": top3,
            "days": days,
        },
    })

@login_required
@require_http_methods(["GET"])
def api_admin_ventas_actividad(request):
    """
    Serie diaria combinada para el dashboard de administrador.

    - ventas: suma diaria de Venta.total (CLP)
    - ordenes: cantidad diaria de compras (Compra)
    - vendedores: cantidad diaria de vendedores con al menos una venta

    Incluye meta.vendedores con los nombres de vendedores por día para tooltips.

    Query params: days (7..365, default 30)
    """

    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("Solo admin")

    try:
        days = int(request.GET.get("days", 30))
    except (TypeError, ValueError):
        days = 30
    days = max(7, min(days, 365))

    hoy = timezone.localdate()
    desde = hoy - timedelta(days=days - 1)

    dias = [desde + timedelta(days=i) for i in range(days)]
    labels = [d.isoformat() for d in dias]

                               
    ventas_qs = (
        Venta.objects
        .filter(fecha_venta__gte=desde, fecha_venta__lte=hoy)
        .values("fecha_venta")
        .annotate(t=Sum("total"))
    )
    ventas_map = {row["fecha_venta"].isoformat(): float(row["t"] or 0) for row in ventas_qs}

                                         
    ordenes_qs = (
        Compra.objects
        .filter(fecha_compra__gte=desde, fecha_compra__lte=hoy)
        .values("fecha_compra")
        .annotate(c=Count("id"))
    )
    ordenes_map = {row["fecha_compra"].isoformat(): int(row["c"] or 0) for row in ordenes_qs}

                                          
    vend_dia_qs = (
        Venta.objects
        .filter(fecha_venta__gte=desde, fecha_venta__lte=hoy)
        .values("fecha_venta", "vendedor__usuario__username")
        .annotate(c=Count("id"))
    )
    vendedores_por_dia = {}
    for row in vend_dia_qs:
        dia = row["fecha_venta"].isoformat()
        nombre = row["vendedor__usuario__username"] or "Vendedor"
        if dia not in vendedores_por_dia:
            vendedores_por_dia[dia] = set()
        vendedores_por_dia[dia].add(nombre)

    ventas_series = [ventas_map.get(lbl, 0.0) for lbl in labels]
    ordenes_series = [ordenes_map.get(lbl, 0) for lbl in labels]
    vendedores_series = [len(sorted(vendedores_por_dia.get(lbl, set()))) for lbl in labels]
    vendedores_meta = [sorted(vendedores_por_dia.get(lbl, set())) for lbl in labels]

    return JsonResponse({
        "labels": labels,
        "series": {
            "ventas": ventas_series,
            "ordenes": ordenes_series,
            "vendedores": vendedores_series,
        },
        "meta": {
            "vendedores": vendedores_meta,
        },
    })
